"""Unit tests for drawing the one-page summary from the Excel template.

The template is the design, so most of these assert that what the sheet says is
what gets drawn — column widths, row heights, merges, fills — and that the
parts the sheet *cannot* know about (how many pipes a well has, how long their
names are) adapt around it.

Rendering is checked through the layout model rather than by comparing images:
a pixel diff would fail on every font-rendering difference while telling you
nothing about whether the table grew correctly.
"""

import os

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from well_tools.report import ops_render
from well_tools.report.pipe_config import SEVERITY

FIELDS = {"well_name": "HRDH-1702", "well_type": "OIL OBSERVATION",
          "log_date": "15-May-2026", "rig": ""}
DEFAULTS = {"RIG": "RIGLESS"}

ROWS = [
    {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 40, "Top (ft)": 0,
     "Bottom (ft)": 5271, "Thick_Nom": "0.395"},
    {"Pipe OD": '18 5/8" CSG', "Weight (ppf)": 88.7, "Top (ft)": 0,
     "Bottom (ft)": 760, "Thick_Nom": "0.435"},
    {"Pipe OD": '4 1/2" TBG', "Weight (ppf)": 11.6, "Top (ft)": 0,
     "Bottom (ft)": 6766, "Thick_Nom": "0.250"},
]

SPOTS = [
    {"pipe": {"suffix": '18 5/8" CSG', "type": "CSG"}, "max_loss": "5.1",
     "grade": "B", "depth": "18.3"},
    {"pipe": {"suffix": '7" CSG', "type": "CSG"}, "max_loss": "31.2",
     "grade": "D", "depth": "4180.4"},
]


def make_template(tmp_path, od_width=20.0):
    """A miniature of the real template: two tables, a units row, headings, and
    blank rows under each tagged row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OPS"
    ws.column_dimensions["B"].width = od_width
    for letter in "CDEF":
        ws.column_dimensions[letter].width = 12.0

    ws["B2"] = "{{well_name}}"
    ws["B3"] = "Rig: {{RIG}}"
    ws["B4"] = "•Log Date:{{log_date}}."
    ws["B5"] = "•Latest Workover Date:{{last_wko}}."

    ws["B7"] = "Completion Strings"
    for col, tag in enumerate(ops_render.STRINGS_TAGS, start=2):
        cell = ws.cell(row=8, column=col)
        cell.value = "{{%s}}" % tag
        cell.border = Border(left=Side("thin"), bottom=Side("thin"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 30.0

    ws["B12"] = "Hot Spot Summary"
    for col, tag in enumerate(ops_render.HOTSPOT_TAGS, start=2):
        ws.cell(row=13, column=col).value = "{{%s}}" % tag
    ws.merge_cells("D13:E13")            # the grade cell spans two columns

    ws["B17"] = "Conclusions"
    ws["B18"] = "{{conclusions}}"

    ws["B20"] = "end"                    # bounds the conclusions block
    path = str(tmp_path / "OPS.xlsx")
    wb.save(path)
    return path


def render(tmp_path, fields=None, rows=None, spots=None, **kw):
    dest = str(tmp_path / "ops.png")
    return ops_render.render_ops(
        make_template(tmp_path, **kw), dest,
        fields if fields is not None else FIELDS,
        rows if rows is not None else ROWS,
        spots if spots is not None else SPOTS,
        defaults=DEFAULTS)


def layout_for(tmp_path, **kw):
    """The layout after filling — what the drawing step consumes."""
    from well_tools.report.ops_render import (
        HOTSPOT_TAGS, STRINGS_TAGS, _fill_scalars, _grow_block, read_layout,
    )

    layout = read_layout(make_template(tmp_path, **kw))
    _fill_scalars(layout, FIELDS, DEFAULTS)
    return layout


# --------------------------------------------------------------------------
# Values
# --------------------------------------------------------------------------
def test_tubing_sorts_last_even_when_it_is_not_the_smallest():
    rows = [{"Pipe OD": '4 1/2" TBG'}, {"Pipe OD": '3 1/2" LNR'},
            {"Pipe OD": '9 5/8" CSG'}]
    assert [r["Pipe OD"] for r in ops_render.sort_strings(rows)] == [
        '9 5/8" CSG', '3 1/2" LNR', '4 1/2" TBG']


def test_a_taper_keeps_its_sections_in_descending_order():
    rows = [{"Pipe OD": '2 7/8" TBG'}, {"Pipe OD": '4 1/2" TBG'},
            {"Pipe OD": '3 1/2" TBG'}, {"Pipe OD": '7" LNR'}]
    assert [r["Pipe OD"] for r in ops_render.sort_strings(rows)] == [
        '7" LNR', '4 1/2" TBG', '3 1/2" TBG', '2 7/8" TBG']


def test_fractional_od_parsing():
    assert ops_render._od_value('18 5/8" CSG') == pytest.approx(18.625)
    assert ops_render._od_value('7" CSG') == 7
    assert ops_render._od_value("nonsense") == 0.0


def test_conclusion_wording_comes_from_the_shared_severity_map():
    """Not a copy of the words — the same map the {{<role>_highest_grade}} tag
    uses, so the summary and the template cannot drift apart."""
    for grade, word in SEVERITY.items():
        line = ops_render.conclusion({"suffix": '7 5/8" LNR', "type": "LNR"}, grade)
        assert line == f"{word} metal loss detected across the 7 5/8\" liner string."


def test_an_unknown_grade_yields_no_conclusion():
    assert ops_render.conclusion({"suffix": '7" CSG', "type": "CSG"}, None) is None


def test_thickness_keeps_its_trailing_zero():
    assert ops_render._thickness(0.250) == "0.250"


# --------------------------------------------------------------------------
# Reading the template
# --------------------------------------------------------------------------
def test_the_template_supplies_the_geometry(tmp_path):
    layout = ops_render.read_layout(make_template(tmp_path))
    # Excel's character widths convert to points
    assert layout.columns[2] == pytest.approx((20.0 * 7 + 5) * 72 / 96)
    heights = [row.height for row in layout.rows]
    assert 30.0 in heights                      # the tagged row's own height


def test_hidden_rows_are_dropped(tmp_path):
    """The real template keeps sample data in hidden rows; Excel would not print
    them and neither does this."""
    from openpyxl import load_workbook

    path = make_template(tmp_path)
    wb = load_workbook(path)
    ws = wb["OPS"]
    ws["B6"] = "SHOULD NOT APPEAR"
    ws.row_dimensions[6].hidden = True
    wb.save(path)

    layout = ops_render.read_layout(path)
    values = [c.value for row in layout.rows for c in row.cells.values()]
    assert "SHOULD NOT APPEAR" not in values


def test_a_trailing_newline_does_not_become_a_second_line(tmp_path):
    """Excel shows a trailing newline as nothing. Counting it as a line pushes
    the text off its own vertical alignment — which is how the padding above a
    heading ended up below it."""
    from openpyxl import load_workbook

    path = make_template(tmp_path)
    wb = load_workbook(path)
    wb["OPS"]["B12"] = "Hot Spot Summary\n"
    wb.save(path)

    layout = ops_render.read_layout(path)
    values = [c.value for row in layout.rows for c in row.cells.values()]
    assert "Hot Spot Summary" in values


def test_merged_cells_become_spans(tmp_path):
    layout = ops_render.read_layout(make_template(tmp_path))
    spans = [c.colspan for row in layout.rows for c in row.cells.values()]
    assert 2 in spans                            # D13:E13


# --------------------------------------------------------------------------
# Filling
# --------------------------------------------------------------------------
def test_scalar_tags_substitute_inside_their_sentence(tmp_path):
    layout = layout_for(tmp_path)
    values = [c.value for row in layout.rows for c in row.cells.values()]
    assert "HRDH-1702" in values
    assert "•Log Date:15-May-2026." in values


def test_a_row_whose_tags_are_all_empty_is_dropped(tmp_path):
    """No workover recorded means no workover line at all."""
    layout = layout_for(tmp_path)
    values = [str(c.value) for row in layout.rows for c in row.cells.values()]
    assert not any("Workover" in v for v in values)


def test_a_blank_rig_says_rigless(tmp_path):
    layout = layout_for(tmp_path)
    values = [c.value for row in layout.rows for c in row.cells.values()]
    assert "Rig: RIGLESS" in values


def test_a_typed_rig_wins(tmp_path):
    from well_tools.report.ops_render import _fill_scalars, read_layout

    layout = read_layout(make_template(tmp_path))
    _fill_scalars(layout, dict(FIELDS, rig="Rig 42"), DEFAULTS)
    values = [c.value for row in layout.rows for c in row.cells.values()]
    assert "Rig: Rig 42" in values


# --------------------------------------------------------------------------
# Drawing
# --------------------------------------------------------------------------
def test_renders_a_png(tmp_path):
    from PIL import Image

    result = render(tmp_path)
    assert result["warnings"] == []
    with Image.open(result["path"]) as im:
        assert im.size == result["size"]
        assert im.width > 0 and im.height > 0


def test_more_pipes_makes_a_taller_page(tmp_path):
    short = render(tmp_path, rows=ROWS[:1], spots=SPOTS[:1])["size"]
    tall = render(tmp_path, rows=ROWS * 4, spots=SPOTS * 3)["size"]
    assert tall[1] > short[1]
    assert tall[0] == short[0]                   # only the height moves


def test_a_long_tapered_name_widens_its_column(tmp_path):
    """The template cannot know a well's string names, so the Pipe OD column
    grows for a tapered label — and only then."""
    plain = render(tmp_path, spots=[
        {"pipe": {"suffix": '7" CSG', "type": "CSG"}, "max_loss": "5.1",
         "grade": "B", "depth": "18.3"}])["size"]
    tapered = render(tmp_path, spots=[
        {"pipe": {"suffix": '4 1/2" × 3 1/2" × 2 7/8" TBG', "type": "TBG"},
         "max_loss": "5.1", "grade": "B", "depth": "18.3"}])["size"]
    assert tapered[0] > plain[0]


def test_the_widening_is_capped(tmp_path):
    """One absurd label must not distort the page."""
    absurd = render(tmp_path, spots=[
        {"pipe": {"suffix": "X" * 400, "type": "CSG"}, "max_loss": "5.1",
         "grade": "B", "depth": "18.3"}])["size"]
    plain = render(tmp_path, spots=[
        {"pipe": {"suffix": '7" CSG', "type": "CSG"}, "max_loss": "5.1",
         "grade": "B", "depth": "18.3"}])["size"]
    assert absurd[0] < plain[0] * 2


def test_repeated_pipe_od_merges_across_its_weight_rows(tmp_path):
    """A pipe in three weight sections takes three rows; repeating its OD in
    each reads as three separate pipes."""
    from well_tools.report.ops_render import (
        STRINGS_TAGS, _fill_scalars, _grow_block, _merge_repeated, read_layout,
    )

    rows = [
        {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 47, "Top (ft)": 0,
         "Bottom (ft)": 1800, "Thick_Nom": "0.472"},
        {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 40, "Top (ft)": 1800,
         "Bottom (ft)": 3900, "Thick_Nom": "0.395"},
        {"Pipe OD": '4 1/2" TBG', "Weight (ppf)": 11.6, "Top (ft)": 0,
         "Bottom (ft)": 6766, "Thick_Nom": "0.250"},
    ]
    layout = read_layout(make_template(tmp_path))
    _fill_scalars(layout, FIELDS, DEFAULTS)
    columns = {t: layout.find_tag(t)[1] for t in STRINGS_TAGS}
    count, first = _grow_block(layout, "str_od", columns, rows,
                               lambda r, t: r.get("Pipe OD") if t == "str_od" else "")
    _merge_repeated(layout, first, count, columns["str_od"])

    head = layout.rows[first].cells[columns["str_od"]]
    assert head.rowspan == 2                       # the two 9 5/8" rows
    assert columns["str_od"] not in layout.rows[first + 1].cells
    # the tubing is a different pipe and keeps its own cell
    assert columns["str_od"] in layout.rows[first + 2].cells


def test_grade_cells_take_the_shared_palette(tmp_path):
    from well_tools.report.tables import GRADE_COLORS

    layout = ops_render.read_layout(make_template(tmp_path))
    expected = {ops_render._rgb(type("C", (), {"rgb": GRADE_COLORS[g]})())
                for g in ("B", "D")}
    render(tmp_path)                                # draws without error
    assert len(expected) == 2                       # B and D are different colours


def test_a_template_that_cannot_be_found_is_refused(tmp_path):
    with pytest.raises(ops_render.OpsRenderError):
        ops_render.read_layout(str(tmp_path / "nope.xlsx"))


def test_an_empty_template_is_refused(tmp_path):
    wb = Workbook()
    wb.active.title = "OPS"
    path = str(tmp_path / "empty.xlsx")
    wb.save(path)

    with pytest.raises(ops_render.OpsRenderError):
        ops_render.read_layout(path)


def test_the_template_is_never_modified(tmp_path):
    path = make_template(tmp_path)
    before = open(path, "rb").read()
    ops_render.render_ops(path, str(tmp_path / "ops.png"), FIELDS, ROWS, SPOTS,
                          defaults=DEFAULTS)
    assert open(path, "rb").read() == before

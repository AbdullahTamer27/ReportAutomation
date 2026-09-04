"""Unit tests for filling the OPS workbook.

The filler writes into a hand-formatted sheet it must not disturb, so most of
what matters here is what survives: a block grows to fit the well, and the
merged ranges, images, heights and formatting below it come along intact.
openpyxl moves cell values on an insert and leaves everything else where it was,
so growing a sheet is only safe because of the repair in `insert_rows` — which
is why several of these tests assert on geometry rather than on values.
"""

import os

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from well_tools.report import ops_fill
from well_tools.report.tables import GRADE_COLORS

FIELDS = {"well_name": "HRDH-1702", "well_type": "OIL OBSERVATION",
          "log_date": "15-May-2026", "rig": ""}

ROWS = [
    {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 40, "Top (ft)": 0,
     "Bottom (ft)": 5271, "Thick_Nom": "0.395"},
    {"Pipe OD": '18 5/8" CSG', "Weight (ppf)": 88.7, "Top (ft)": 0,
     "Bottom (ft)": 760, "Thick_Nom": "0.435"},
    {"Pipe OD": '4 1/2" TBG', "Weight (ppf)": 11.6, "Top (ft)": 0,
     "Bottom (ft)": 6766, "Thick_Nom": "0.250"},
]

SPOTS = [
    {"pipe": {"suffix": '18 5/8" CSG', "type": "CSG"}, "max_loss": "5.1",
     "grade": "B", "depth": "18.3"},
    {"pipe": {"suffix": '7" CSG', "type": "CSG"}, "max_loss": "31.2",
     "grade": "D", "depth": "4180.4"},
]

DEFAULTS = {"RIG": "Rigless"}


def make_template(tmp_path, blank_rows=6):
    """A miniature of the real template: tagged rows, blank rows beneath them,
    and a heading right after — which is what bounds each block."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OPS"
    ws["B2"] = "{{well_name}}"
    ws["B3"] = "Rig: {{RIG}}"
    ws["B4"] = "•Log Date:{{log_date}}."
    ws["B5"] = "•Latest Workover Date:{{last_wko}}."

    row = 7
    ws.cell(row=row, column=2).value = "Completion Strings"
    row += 1
    for col, tag in enumerate(ops_fill.STRINGS_TAGS, start=2):
        ws.cell(row=row, column=col).value = "{{%s}}" % tag
    strings_row = row
    row += 1 + blank_rows

    ws.cell(row=row, column=2).value = "Hot Spot Summary"
    row += 1
    for col, tag in enumerate(ops_fill.HOTSPOT_TAGS, start=2):
        ws.cell(row=row, column=col).value = "{{%s}}" % tag
    hotspot_row = row
    row += 1 + blank_rows

    ws.cell(row=row, column=2).value = "Conclusions"
    row += 1
    ws.cell(row=row, column=2).value = "{{conclusions}}"
    conclusions_row = row
    for extra in range(1, blank_rows + 1):
        ws.cell(row=row + extra, column=2).value = None

    path = str(tmp_path / "OPS.xlsx")
    wb.save(path)
    return path, strings_row, hotspot_row, conclusions_row


def fill(tmp_path, fields=None, rows=None, spots=None, blank_rows=6):
    template, s_row, h_row, c_row = make_template(tmp_path, blank_rows)
    dest = str(tmp_path / "out.xlsx")
    result = ops_fill.fill_ops(template, dest,
                               fields if fields is not None else FIELDS,
                               rows if rows is not None else ROWS,
                               spots if spots is not None else SPOTS,
                               defaults=DEFAULTS)
    return load_workbook(dest)["OPS"], result, (s_row, h_row, c_row)


# --------------------------------------------------------------------------
# Sorting
# --------------------------------------------------------------------------
def test_tubing_sorts_last_even_when_it_is_not_the_smallest():
    """Tubing is the innermost string whatever its diameter, so a 3 1/2" liner
    still belongs above a 4 1/2" tubing — which sorting on OD alone reverses."""
    rows = [{"Pipe OD": '4 1/2" TBG'}, {"Pipe OD": '3 1/2" LNR'},
            {"Pipe OD": '9 5/8" CSG'}]
    assert [r["Pipe OD"] for r in ops_fill.sort_strings(rows)] == [
        '9 5/8" CSG', '3 1/2" LNR', '4 1/2" TBG']


def test_casings_and_liners_sort_together_by_od():
    rows = [{"Pipe OD": '7" CSG'}, {"Pipe OD": '7 5/8" LNR'},
            {"Pipe OD": '18 5/8" CSG'}]
    assert [r["Pipe OD"] for r in ops_fill.sort_strings(rows)] == [
        '18 5/8" CSG', '7 5/8" LNR', '7" CSG']


def test_fractional_od_parsing():
    assert ops_fill._od_value('18 5/8" CSG') == pytest.approx(18.625)
    assert ops_fill._od_value('7" CSG') == 7
    assert ops_fill._od_value("nonsense") == 0.0


# --------------------------------------------------------------------------
# Scalars and defaults
# --------------------------------------------------------------------------
def test_tags_substitute_inside_their_sentence(tmp_path):
    ws, _, _ = fill(tmp_path)
    assert ws["B2"].value == "HRDH-1702"
    assert ws["B4"].value == "•Log Date:15-May-2026."


def test_a_row_whose_tags_are_all_empty_is_hidden(tmp_path):
    """No workover recorded means no workover line — printing "Date: ." would
    read as a finding rather than a blank."""
    ws, _, _ = fill(tmp_path)
    assert ws.row_dimensions[5].hidden is True
    assert ws["B5"].value is None


def test_a_blank_rig_says_rigless(tmp_path):
    ws, _, _ = fill(tmp_path)
    assert ws["B3"].value == "Rig: Rigless"
    assert ws.row_dimensions[3].hidden is False


def test_na_counts_as_blank_for_the_rig(tmp_path):
    ws, _, _ = fill(tmp_path, fields=dict(FIELDS, rig="N/A"))
    assert ws["B3"].value == "Rig: Rigless"


def test_a_typed_rig_is_written_through(tmp_path):
    """The tag is {{RIG}} but the payload key is "rig" — matched without case,
    or every rig anyone types would silently come out as "Rigless"."""
    ws, _, _ = fill(tmp_path, fields=dict(FIELDS, rig="Rig 42"))
    assert ws["B3"].value == "Rig: Rig 42"


# --------------------------------------------------------------------------
# Blocks
# --------------------------------------------------------------------------
def test_rows_are_written_in_order_and_leftovers_hidden(tmp_path):
    ws, result, (s_row, _, _) = fill(tmp_path)
    assert result["warnings"] == []
    assert [ws.cell(row=s_row + i, column=2).value for i in range(3)] == [
        '18 5/8" CSG', '9 5/8" CSG', '4 1/2" TBG']
    # the 4 unused rows of the block are hidden and cleared
    for offset in range(3, 7):
        assert ws.row_dimensions[s_row + offset].hidden is True
        assert ws.cell(row=s_row + offset, column=2).value is None


def test_the_block_grows_past_its_blank_rows(tmp_path):
    """More data than blank rows: the block inserts what it needs and pushes the
    next section down, rather than overwriting it or dropping rows."""
    rows = ROWS * 4                                   # 12 rows into room for 3
    ws, result, (s_row, _, _) = fill(tmp_path, rows=rows, blank_rows=2)

    assert result["warnings"] == []
    # All 12 rows are present. Asserted on a column that isn't merged: the Pipe
    # OD column deliberately merges runs of the same pipe, so only the first row
    # of each run still carries its label.
    bottoms = [ws.cell(row=s_row + i, column=5).value for i in range(12)]
    assert bottoms == [float(r["Bottom (ft)"]) for r in ops_fill.sort_strings(rows)]
    # the heading was pushed below the grown block, not written over
    assert ws.cell(row=s_row + 12, column=2).value == "Hot Spot Summary"


def test_a_template_with_no_blank_rows_at_all_still_works(tmp_path):
    """The point of growing: one tagged row is a complete template. Nothing has
    to be reserved, and no count here has to match one over there."""
    ws, result, (s_row, _, _) = fill(tmp_path, blank_rows=0)

    assert result["warnings"] == []
    assert [ws.cell(row=s_row + i, column=2).value for i in range(3)] == [
        '18 5/8" CSG', '9 5/8" CSG', '4 1/2" TBG']
    assert ws.cell(row=s_row + 3, column=2).value == "Hot Spot Summary"


def test_sections_keep_their_order_and_content(tmp_path):
    """Blocks grow, so rows move — but every heading must still be there, in
    order, with nothing written over it."""
    template, s_row, h_row, c_row = make_template(tmp_path)
    before = load_workbook(template)["OPS"]
    headings = {ws_row: before.cell(row=ws_row, column=2).value
                for ws_row in range(1, before.max_row + 1)
                if before.cell(row=ws_row, column=2).value in
                ("Completion Strings", "Hot Spot Summary", "Conclusions")}

    dest = str(tmp_path / "out.xlsx")
    ops_fill.fill_ops(template, dest, FIELDS, ROWS, SPOTS, defaults=DEFAULTS)
    after = load_workbook(dest)["OPS"]

    for row, text in headings.items():
        assert after.cell(row=row, column=2).value == text


def test_grade_cells_take_the_shared_palette(tmp_path):
    ws, _, (_, h_row, _) = fill(tmp_path)
    grade_col = 2 + ops_fill.HOTSPOT_TAGS.index("hs_grade")
    for offset, spot in enumerate(SPOTS):
        fill_colour = ws.cell(row=h_row + offset, column=grade_col).fill.fgColor.rgb
        assert GRADE_COLORS[spot["grade"]] in str(fill_colour)


def test_max_wl_is_written_as_a_fraction_for_the_percent_format(tmp_path):
    """The template's Max WL cell is percent-formatted, which expects 0.112 for
    11.2% — the engine carries the value in percent units."""
    ws, _, (_, h_row, _) = fill(tmp_path)
    assert ws.cell(row=h_row, column=3).value == pytest.approx(0.051)


def test_conclusions_come_from_the_hot_spots_plus_the_boilerplate(tmp_path):
    ws, _, (_, _, c_row) = fill(tmp_path)
    lines = [ws.cell(row=c_row + i, column=2).value for i in range(3)]
    assert lines[0].startswith("Minor metal loss detected across the 18 5/8")
    assert lines[1].startswith("Intensive metal loss detected across the 7")
    assert lines[2] == ops_fill.TEMPERATURE_NOTE


def test_repeated_pipe_od_is_merged_across_its_weight_rows(tmp_path):
    """A pipe with three weight sections takes three rows; leaving its OD
    repeated in each reads as three separate pipes."""
    rows = [
        {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 40, "Top (ft)": 0,
         "Bottom (ft)": 100, "Thick_Nom": "0.395"},
        {"Pipe OD": '9 5/8" CSG', "Weight (ppf)": 47, "Top (ft)": 100,
         "Bottom (ft)": 900, "Thick_Nom": "0.472"},
        {"Pipe OD": '4 1/2" TBG', "Weight (ppf)": 11.6, "Top (ft)": 0,
         "Bottom (ft)": 900, "Thick_Nom": "0.250"},
    ]
    ws, _, (s_row, _, _) = fill(tmp_path, rows=rows)
    merged = [str(m) for m in ws.merged_cells.ranges]
    assert any(m.startswith("B%d:B%d" % (s_row, s_row + 1)) for m in merged), merged


def test_the_template_is_never_modified(tmp_path):
    template, _, _, _ = make_template(tmp_path)
    before = open(template, "rb").read()
    ops_fill.fill_ops(template, str(tmp_path / "out.xlsx"), FIELDS, ROWS, SPOTS,
                      defaults=DEFAULTS)
    assert open(template, "rb").read() == before


def test_a_template_without_the_required_tags_is_refused(tmp_path):
    wb = Workbook()
    wb.active["A1"] = "no tags here"
    path = str(tmp_path / "bare.xlsx")
    wb.save(path)

    with pytest.raises(ops_fill.OpsFillError):
        ops_fill.fill_ops(path, str(tmp_path / "out.xlsx"), FIELDS, ROWS, SPOTS)


def test_growing_carries_merges_images_and_heights_with_it(tmp_path):
    """The bug that made insertion look impossible: openpyxl moves cell values on
    an insert and leaves merged ranges behind. Shifting the ranges in place — not
    unmerging and remerging, which blanks cells — is what makes this safe."""
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image

    template, s_row, h_row, c_row = make_template(tmp_path, blank_rows=1)
    wb = load_workbook(template)
    ws = wb["OPS"]
    # a merged cell inside the hot-spot tagged row, as the real template has
    grade_col = 2 + ops_fill.HOTSPOT_TAGS.index("hs_grade")
    ws.merge_cells(start_row=h_row, start_column=grade_col,
                   end_row=h_row, end_column=grade_col + 1)
    ws.row_dimensions[s_row].height = 39
    Image.new("RGB", (40, 20), (0, 0, 0)).save(str(tmp_path / "logo.png"))
    ws.add_image(XLImage(str(tmp_path / "logo.png")), "H%d" % (c_row + 2))
    wb.save(template)

    dest = str(tmp_path / "out.xlsx")
    ops_fill.fill_ops(template, dest, FIELDS, ROWS, SPOTS, defaults=DEFAULTS)
    out = load_workbook(dest)["OPS"]

    # every hot-spot row kept the merge, wherever it ended up. Located from the
    # heading, since the same pipe label also appears in the strings table above.
    heading = next(c.row for c in out["B"] if c.value == "Hot Spot Summary")
    hs_rows = [heading + 1 + i for i in range(len(SPOTS))]
    merged = {(m.min_row, m.min_col) for m in out.merged_cells.ranges}
    for row in hs_rows:
        assert (row, grade_col) in merged, f"row {row} lost its grade merge"

    # values survived the shift — this is what remerging destroyed
    for row, spot in zip(hs_rows, SPOTS):
        assert out.cell(row=row, column=2).value == spot["pipe"]["suffix"]
        assert out.cell(row=row, column=grade_col).value == spot["grade"]

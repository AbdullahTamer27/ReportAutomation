"""Tab 1 — Interval Generator.

This is the existing, working tool. Logic is unchanged: it has only been
moved into a notebook tab (takes a `parent` frame instead of the root window)
and now imports its helpers from the `core` package. Do not change behavior here.
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from .dnd import DND_FILES
from ..core.xml_parser import parse_wellschematic_xml, build_pipe_summary
from ..core.intervals import build_intervals_from_xml
from ..core.thickness import parse_thickness_sections
from ..core.excel_output import (
    write_raw_data_to_template,
    write_raw_data_to_new_file,
)


class IntervalTab:
    def __init__(self, parent, dnd_enabled):
        self.parent = parent
        self.dnd_enabled = dnd_enabled

        self.xml_data = None
        self.xml_file_path = None
        self.template_file_path = None
        self.thickness_path = None
        self.thickness_sections = None

        self.container = ttk.Frame(parent)
        self.container.pack(fill="both", expand=True, padx=10, pady=10)

        self.setup_ui()

    def setup_ui(self):
        info_frame = ttk.LabelFrame(self.container, text="Instructions")
        info_frame.pack(fill="x", padx=10, pady=5)
        instructions = ("1. Load a WellSchematic.xml file (or drag & drop)\n"
                        "2. (OPTIONAL) Load your Excel template — if provided, the file is updated\n"
                        "   in place; if skipped, a brand-new Excel file is created. If the template\n"
                        "   contains a sheet named 'THICKNESS', each interval also gets Channel and\n"
                        "   Offset rows (mode value per pipe). No THICKNESS sheet → those rows are omitted.\n"
                        "3. Click 'Generate Raw Data' — the 'Raw Data' sheet is created or\n"
                        "   overwritten. Other sheets in the template are not touched.")
        ttk.Label(info_frame, text=instructions, justify="left").pack(padx=10, pady=10)

        xml_frame = ttk.LabelFrame(self.container, text="Step 1 — XML Well Schematic")
        xml_frame.pack(fill="x", padx=10, pady=5)
        self.xml_info_label = ttk.Label(xml_frame, text="No XML file loaded",
                                         foreground="gray", font=("Arial", 10))
        self.xml_info_label.pack(pady=5)
        ttk.Button(xml_frame, text="Load WellSchematic.xml",
                   command=self.load_xml_for_intervals).pack(pady=5)

        template_frame = ttk.LabelFrame(self.container, text="Step 2 — Excel Template (Optional)")
        template_frame.pack(fill="x", padx=10, pady=5)
        self.template_info_label = ttk.Label(template_frame,
                                             text="No template loaded — a new file will be created",
                                             foreground="gray", font=("Arial", 10))
        self.template_info_label.pack(pady=5)
        template_btn_frame = ttk.Frame(template_frame)
        template_btn_frame.pack(pady=5)
        ttk.Button(template_btn_frame, text="Load Excel Template",
                   command=self.load_template).pack(side="left", padx=5)
        ttk.Button(template_btn_frame, text="Clear Template",
                   command=self.clear_template).pack(side="left", padx=5)

        preview_frame = ttk.LabelFrame(self.container, text="Preview")
        preview_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.xml_preview = tk.Text(preview_frame, height=15, wrap="word", font=("Courier", 9))
        self.xml_preview.pack(fill="both", expand=True)
        preview_scroll = ttk.Scrollbar(preview_frame, command=self.xml_preview.yview)
        preview_scroll.pack(side="right", fill="y")
        self.xml_preview.config(yscrollcommand=preview_scroll.set)

        ttk.Button(self.container, text="Step 3 — Generate Raw Data",
                   command=self.export_raw_data).pack(pady=10)

        if self.dnd_enabled:
            self._setup_drag_and_drop(xml_frame, template_frame)
            self.xml_info_label.config(text="No XML file loaded  (or drag & drop here)")
            self.template_info_label.config(
                text="No template loaded — a new file will be created  (or drag & drop)")

    # ---- Drag & drop ----

    def _parse_drop_data(self, data):
        paths = []
        current = ""
        in_brace = False
        for ch in data:
            if ch == '{':
                in_brace = True
            elif ch == '}':
                in_brace = False
                if current:
                    paths.append(current)
                    current = ""
            elif ch == ' ' and not in_brace:
                if current:
                    paths.append(current)
                    current = ""
            else:
                current += ch
        if current:
            paths.append(current)
        return paths

    def _setup_drag_and_drop(self, xml_frame, template_frame):
        for widget in (xml_frame, template_frame, self.container):
            widget.drop_target_register(DND_FILES)
        xml_frame.dnd_bind('<<Drop>>', self._on_drop_xml)
        template_frame.dnd_bind('<<Drop>>', self._on_drop_template)
        self.container.dnd_bind('<<Drop>>', self._on_drop_smart)

    def _on_drop_xml(self, event):
        paths = self._parse_drop_data(event.data)
        if not paths:
            return
        path = paths[0]
        if not path.lower().endswith('.xml'):
            messagebox.showwarning("Wrong File Type",
                f"Expected an XML file, got:\n{os.path.basename(path)}")
            return
        self.load_xml_for_intervals(path)

    def _on_drop_template(self, event):
        paths = self._parse_drop_data(event.data)
        if not paths:
            return
        path = paths[0]
        if not path.lower().endswith(('.xlsx', '.xlsm')):
            messagebox.showwarning("Wrong File Type",
                f"Expected an Excel file (.xlsx or .xlsm), got:\n{os.path.basename(path)}")
            return
        self.load_template(path)

    def _on_drop_smart(self, event):
        paths = self._parse_drop_data(event.data)
        for path in paths:
            ext = path.lower()
            if ext.endswith('.xml'):
                self.load_xml_for_intervals(path)
            elif ext.endswith(('.xlsx', '.xlsm')):
                self.load_template(path)

    # ---- Loading ----

    def load_xml_for_intervals(self, file_path=None):
        if file_path is None:
            file_path = filedialog.askopenfilename(
                title="Select WellSchematic XML File",
                filetypes=[("XML Files", "*.xml"), ("All Files", "*.*")]
            )
        if not file_path:
            return
        try:
            self.xml_data = parse_wellschematic_xml(file_path)
            self.xml_file_path = file_path
            num_pipes = len(self.xml_data)
            pipe_types = self.xml_data['Type'].value_counts().to_dict()
            type_str = ", ".join([f"{count} {typ}" for typ, count in pipe_types.items()])
            info_text = f"✓ Loaded: {num_pipes} pipes ({type_str})\n"
            info_text += f"Depth range: {self.xml_data['Start'].min():.0f} - {self.xml_data['End'].max():.0f} ft"
            self.xml_info_label.config(text=info_text, foreground="green")
            self.xml_preview.delete("1.0", tk.END)
            self.xml_preview.insert(tk.END, "===== XML Data Loaded =====\n\n")
            preview_df = self.xml_data[['Type', 'OD', 'ID', 'Weight', 'Thickness', 'Start', 'End']]
            self.xml_preview.insert(tk.END, preview_df.to_string(index=False))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse XML:\n{str(e)}")
            self.xml_info_label.config(text="✗ Error loading XML", foreground="red")

    def load_template(self, file_path=None):
        from openpyxl import load_workbook
        if file_path is None:
            file_path = filedialog.askopenfilename(
                title="Select Excel Template",
                filetypes=[
                    ("Excel Files", "*.xlsx *.xlsm"),
                    ("Excel Macro-Enabled", "*.xlsm"),
                    ("Excel Workbook", "*.xlsx"),
                    ("All Files", "*.*")
                ]
            )
        if not file_path:
            return
        try:
            is_macro = file_path.lower().endswith('.xlsm')
            wb = load_workbook(file_path, keep_vba=is_macro)
            sheet_names = wb.sheetnames
            wb.close()
            self.template_file_path = file_path
            macro_tag = " [macro-enabled]" if is_macro else ""
            info_text = f"✓ Template: {os.path.basename(file_path)}{macro_tag}\n"
            info_text += f"Existing sheets: {', '.join(sheet_names)}"

            self.thickness_path = None
            self.thickness_sections = None
            channel_note = ""
            try:
                sections = parse_thickness_sections(file_path)
                if sections:
                    self.thickness_path = file_path
                    self.thickness_sections = sections
                    channel_note = (f"\nTHICKNESS sheet found ({len(sections)} sections) "
                                    f"— Channel/Offset rows will be added.")
                else:
                    channel_note = ("\nTHICKNESS sheet present but unreadable "
                                    "— Channel/Offset rows omitted.")
            except ValueError as te:
                if str(te) == "NO_THICKNESS_SHEET":
                    channel_note = "\nNo THICKNESS sheet — Channel/Offset rows omitted."
                else:
                    channel_note = f"\nTHICKNESS read issue: {te} — rows omitted."

            self.template_info_label.config(text=info_text + channel_note, foreground="green")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open template:\n{str(e)}")
            self.template_info_label.config(text="✗ Error loading template", foreground="red")

    def clear_template(self):
        self.template_file_path = None
        self.thickness_path = None
        self.thickness_sections = None
        if self.dnd_enabled:
            self.template_info_label.config(
                text="No template loaded — a new file will be created  (or drag & drop)",
                foreground="gray")
        else:
            self.template_info_label.config(
                text="No template loaded — a new file will be created",
                foreground="gray")

    # ---- Export ----

    def export_raw_data(self):
        if self.xml_data is None:
            messagebox.showwarning("No XML", "Please load an XML file first (Step 1)")
            return
        try:
            pipe_summary_df = build_pipe_summary(self.xml_data)
            interval_df = build_intervals_from_xml(
                self.xml_data, thickness_sections=self.thickness_sections)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to build tables from XML:\n{str(e)}")
            return
        if self.template_file_path is None:
            self._export_to_new_file(pipe_summary_df, interval_df)
        else:
            self._export_to_template(pipe_summary_df, interval_df)

    def _export_to_new_file(self, pipe_summary_df, interval_df):
        xml_basename = os.path.splitext(os.path.basename(self.xml_file_path or ""))[0]
        default_name = f"{xml_basename or 'WellSchematic'}_RawData.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Save Raw Data Excel File",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if not output_path:
            return
        if not output_path.lower().endswith('.xlsx'):
            output_path = os.path.splitext(output_path)[0] + '.xlsx'
        try:
            write_raw_data_to_new_file(output_path, pipe_summary_df, interval_df)
            self._show_export_preview(pipe_summary_df, interval_df, output_path,
                                      action_label="Created new file")
            messagebox.showinfo("Success",
                f"New Excel file created:\n{os.path.basename(output_path)}")
        except PermissionError:
            messagebox.showerror("File Locked",
                "Can't write to that location — the file may be open in Excel.\n\n"
                "Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create file:\n{str(e)}")

    def _export_to_template(self, pipe_summary_df, interval_df):
        template_name = os.path.basename(self.template_file_path)
        confirm = messagebox.askyesno(
            "Confirm Update",
            f"This will update the 'Raw Data' sheet in:\n\n{template_name}\n\n"
            f"If the sheet doesn't exist it will be created.\n"
            f"If it does, it will be wiped and rewritten.\n\n"
            f"The rest of the workbook is untouched. Proceed?"
        )
        if not confirm:
            return
        try:
            write_raw_data_to_template(
                self.template_file_path, pipe_summary_df, interval_df)
            self._show_export_preview(pipe_summary_df, interval_df,
                                      self.template_file_path,
                                      action_label="Updated in place")
            messagebox.showinfo("Success",
                f"'Raw Data' sheet updated successfully in:\n{template_name}")
        except PermissionError:
            messagebox.showerror("File Locked",
                "Can't write to the template — it may be open in Excel.\n\n"
                "Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update template:\n{str(e)}")

    def _show_export_preview(self, pipe_summary_df, interval_df, output_path, action_label):
        self.xml_preview.delete("1.0", tk.END)
        self.xml_preview.insert(tk.END, "===== TABLE 1 — INTERVALS =====\n\n")
        for i, row in interval_df.iterrows():
            self.xml_preview.insert(tk.END,
                f"Interval {i+1}: {row['Start Depth (ft)']} – {row['End Depth (ft)']} ft\n")
            for cfg in row["Configurations"]:
                self.xml_preview.insert(tk.END, f"   • {cfg}\n")
            if "Channels" in interval_df.columns:
                self.xml_preview.insert(tk.END,
                    f"   Channel: {'-'.join(str(v) for v in row['Channels'])}\n")
                self.xml_preview.insert(tk.END,
                    f"   Offset:  {'/'.join(str(v) for v in row['Offsets'])}\n")
            self.xml_preview.insert(tk.END, "\n")
        self.xml_preview.insert(tk.END, "===== TABLE 2 — PIPE SUMMARY =====\n\n")
        self.xml_preview.insert(tk.END, pipe_summary_df.to_string(index=False))
        self.xml_preview.insert(tk.END, f"\n\n✅ {action_label}:\n{output_path}")

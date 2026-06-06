"""Parsing of the optional THICKNESS sheet and per-pipe channel/offset resolution."""

import re
from openpyxl import load_workbook

from .formatting import format_weight

OD_MATCH_TOL = 0.01          # ODs equal if within this many inches
NA_TEXT = "N/A"              # shown when no channel/offset matches a pipe
THICKNESS_SHEET_NAME = "THICKNESS"


def _is_na(v):
    return v is None or (isinstance(v, str) and v.strip().upper() in ("", "N/A"))


def _format_channel(v):
    if _is_na(v):
        return None
    s = str(v).strip()
    m = re.match(r'^[A-Za-z]*([0-9].*)$', s)
    return m.group(1) if m else s


def _format_offset(v):
    if _is_na(v):
        return None
    try:
        return format_weight(round(float(v), 3))
    except (TypeError, ValueError):
        return str(v).strip()


def parse_thickness_sections(xlsx_path):
    is_macro = xlsx_path.lower().endswith('.xlsm')
    wb = load_workbook(xlsx_path, data_only=True, keep_vba=is_macro, read_only=True)
    try:
        # Case-insensitive sheet lookup.
        match = next((s for s in wb.sheetnames
                      if s.strip().upper() == THICKNESS_SHEET_NAME), None)
        if match is None:
            raise ValueError("NO_THICKNESS_SHEET")
        ws = wb[match]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def find_col(name):
        for i, h in enumerate(header):
            if h.upper() == name.upper():
                return i
        return None

    top_c = find_col("Top Depth")
    bot_c = find_col("Bottom Depth")
    if top_c is None or bot_c is None:
        raise ValueError("THICKNESS sheet missing 'Top Depth'/'Bottom Depth' columns")

    slot_cols = []
    i = 1
    while True:
        od_c = find_col(f"OD{i}")
        if od_c is None:
            break
        slot_cols.append({
            "od": od_c,
            "channel": find_col(f"P{i}ThkCh"),
            "offset": find_col(f"P{i}Offset"),
        })
        i += 1

    sections = []
    for r in rows[1:]:
        if top_c >= len(r) or bot_c >= len(r):
            continue
        if r[top_c] is None or r[bot_c] is None:
            continue
        try:
            top, bottom = float(r[top_c]), float(r[bot_c])
        except (TypeError, ValueError):
            continue

        slots = []
        for sc in slot_cols:
            od_idx = sc["od"]
            if od_idx is None or od_idx >= len(r) or _is_na(r[od_idx]):
                continue
            try:
                od = float(r[od_idx])
            except (TypeError, ValueError):
                continue
            ch = r[sc["channel"]] if sc["channel"] is not None and sc["channel"] < len(r) else None
            off = r[sc["offset"]] if sc["offset"] is not None and sc["offset"] < len(r) else None
            slots.append({"od": od, "channel": ch, "offset": off})

        if slots:
            sections.append({"top": top, "bottom": bottom, "slots": slots})

    return sections


def _mode_for_pipe(sections, od, start, end, field, formatter):
    candidates = []  # (formatted_value, overlap_footage)
    for s in sections:
        overlap = min(end, s["bottom"]) - max(start, s["top"])
        if overlap <= 0:
            continue
        for slot in s["slots"]:
            if abs(slot["od"] - od) <= OD_MATCH_TOL:
                val = formatter(slot[field])
                if val is not None:
                    candidates.append((val, overlap))

    if not candidates:
        return None

    counts, footage = {}, {}
    order = {}
    for idx, (val, foot) in enumerate(candidates):
        counts[val] = counts.get(val, 0) + 1
        footage[val] = footage.get(val, 0.0) + foot
        order.setdefault(val, idx)
    # Highest count, then most footage, then earliest appearance.
    return max(counts, key=lambda v: (counts[v], footage[v], -order[v]))

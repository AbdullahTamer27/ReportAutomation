"""Writes the 'Raw Data' sheet (Table 1 intervals + Table 2 pipe summary)."""

import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _build_raw_data_sheet(ws, pipe_summary_df, interval_df):
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    label_font  = Font(bold=True, size=10)
    label_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    data_font   = Font(size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin = Side(style='thin', color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_intervals = len(interval_df)
    max_tubulars = max((len(c) for c in interval_df["Configurations"]), default=0)

    ws.cell(row=1, column=1).value = "TABLE 1 — INTERVALS"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="366092")

    header_row = 2
    corner = ws.cell(row=header_row, column=1)
    corner.fill = header_fill
    corner.border = border

    for i in range(num_intervals):
        c = ws.cell(row=header_row, column=2 + i)
        c.value = f"Interval {i + 1}"
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    attribute_rows = ["Start Depth (ft)", "End Depth (ft)"]
    for i in range(max_tubulars):
        attribute_rows.append(f"Tubular {i + 1}")

    for row_offset, label in enumerate(attribute_rows, start=1):
        r = header_row + row_offset
        lc = ws.cell(row=r, column=1)
        lc.value = label
        lc.font = label_font
        lc.fill = label_fill
        lc.alignment = left
        lc.border = border

        for i, (_, interval) in enumerate(interval_df.iterrows()):
            c = ws.cell(row=r, column=2 + i)
            c.font = data_font
            c.alignment = center
            c.border = border
            if label == "Start Depth (ft)":
                c.value = interval["Start Depth (ft)"]
            elif label == "End Depth (ft)":
                c.value = interval["End Depth (ft)"]
            else:
                tub_idx = row_offset - 3
                configs = interval["Configurations"]
                if tub_idx < len(configs):
                    c.value = configs[tub_idx]
                else:
                    c.value = "/"

    last_attr_row = header_row + len(attribute_rows)
    has_channels = "Channels" in interval_df.columns
    if has_channels:
        extra = [("Channel", "Channels", "-"), ("Offset", "Offsets", "/")]
        for k, (label, col, sep) in enumerate(extra, start=1):
            r = last_attr_row + k
            lc = ws.cell(row=r, column=1)
            lc.value = label
            lc.font = label_font
            lc.fill = label_fill
            lc.alignment = left
            lc.border = border
            for i, (_, interval) in enumerate(interval_df.iterrows()):
                c = ws.cell(row=r, column=2 + i)
                c.font = data_font
                c.alignment = center
                c.border = border
                vals = interval[col]
                c.value = sep.join(str(v) for v in vals) if vals else "/"
        last_attr_row += len(extra)

    table1_last_row = last_attr_row
    t2_start_row = table1_last_row + 3

    ws.cell(row=t2_start_row, column=1).value = "TABLE 2 — PIPE SUMMARY"
    ws.cell(row=t2_start_row, column=1).font = Font(bold=True, size=12, color="366092")
    ws.merge_cells(start_row=t2_start_row, start_column=1,
                   end_row=t2_start_row, end_column=5)

    t2_header_row = t2_start_row + 1
    t2_headers = ["Pipe OD", "Weight (ppf)", "Top (ft)", "Bottom (ft)", "Thick_Nom"]
    for col_idx, h in enumerate(t2_headers, start=1):
        c = ws.cell(row=t2_header_row, column=col_idx)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, row in pipe_summary_df.iterrows():
        r = t2_header_row + 1 + i
        for col_idx, h in enumerate(t2_headers, start=1):
            c = ws.cell(row=r, column=col_idx)
            c.value = row[h]
            c.font = data_font
            c.alignment = center if col_idx > 1 else left
            c.border = border

    ws.column_dimensions['A'].width = 22
    for i in range(num_intervals):
        col_letter = get_column_letter(2 + i)
        ws.column_dimensions[col_letter].width = 28


def write_raw_data_to_template(template_path, pipe_summary_df, interval_df):
    template_dir = os.path.dirname(os.path.abspath(template_path))
    template_ext = os.path.splitext(template_path)[1]
    temp_path = os.path.join(template_dir, f".~tmp_rawdata_{os.getpid()}{template_ext}")

    try:
        shutil.copy(template_path, temp_path)
        is_macro_enabled = template_path.lower().endswith('.xlsm')
        wb = load_workbook(temp_path, keep_vba=is_macro_enabled)
        if "Raw Data" in wb.sheetnames:
            del wb["Raw Data"]
        ws = wb.create_sheet("Raw Data")
        _build_raw_data_sheet(ws, pipe_summary_df, interval_df)
        wb.save(temp_path)
        wb.close()
        os.replace(temp_path, template_path)
    except Exception:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise
    return template_path


def write_raw_data_to_new_file(output_path, pipe_summary_df, interval_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    _build_raw_data_sheet(ws, pipe_summary_df, interval_df)
    wb.save(output_path)
    return output_path

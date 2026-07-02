"""Floating-overlay text boxes — a deliberately self-contained pass.

Overlays are anchored text boxes whose visible text is an ``{{ovl_...}}`` tag.
This module fills those tags and removes the boxes that don't apply. It lives on
its own on purpose: it imports nothing from the rest of the report engine and
shares no helpers with it. Overlay tags only ever appear *inside text boxes*
(``w:txbxContent``) — a region no other pass touches — so a bug here cannot reach
the working code, and a working-code bug cannot reach here. If something is wrong
with an overlay, this is the only file to read.

Implemented:
    {{ovl_wellhead}}        -> well-head damage / clean statement (a checkbox picks).
    {{ovl_shoe_<role>}}     -> "<size> <Type> Shoe at <depth>ft" for each pipe,
                               except the bottom (at/below logged TD) shoe.
    {{ovl_hanger_<role>}}   -> "<size> Liner Hanger at <depth>ft" for each liner.

`<role>` is firstPipe … seventhPipe. Boxes whose tag isn't filled (absent pipe,
omitted bottom shoe, non-liner hanger) are removed entirely.
"""

import re

from docx import Document
from docx.oxml.ns import qn

# --- Well-head overlay text (overlay #1) -------------------------------------
WELLHEAD_TAG = "{{ovl_wellhead}}"
WELLHEAD_DAMAGE = (
    "Pipes’ damage or well-head effect observed below well-head interval."
)
WELLHEAD_CLEAN = (
    "No clear pipes’ damage  or well-head effect observed "
    "around well-head interval."
)

# Inches in callouts use the typographic right double-quote, e.g. 18 5/8”.
_INCH = "”"
# A shoe within this many feet of the deepest shoe is the bottom string → omitted.
_TD_TOL = 1.0

_W_T = qn("w:t")
_W_P = qn("w:p")
_W_R = qn("w:r")
_W_PPR = qn("w:pPr")
_W_TXBX = qn("w:txbxContent")

# Variable overlay tags removed when not filled (never the fixed wellhead):
# shoe/hanger callouts, and the per-damage-point metal-loss / channel slots.
_REMOVABLE = re.compile(
    r"\{\{ovl_(?:shoe_\w+|hanger_\w+|ml\d+_\d+|ch\d+_\d+)\}\}"
)


# --- Local helpers (intentionally not shared) --------------------------------
def _format_depth(value):
    """Depth to at most one decimal, dropping a trailing '.0' (444.6, 1201)."""
    if value is None:
        return ""
    s = f"{float(value):.1f}"
    return s[:-2] if s.endswith(".0") else s


def _load_wb(excel_path):
    if not excel_path:
        return None
    try:
        import openpyxl
        return openpyxl.load_workbook(excel_path, data_only=True)
    except Exception:  # noqa: BLE001
        return None


def _min_top_body(wb, sheet):
    """Liner hanger depth = shallowest Top Body (col B) of the sheet's joints."""
    if not wb or not sheet or sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    tops = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(row=r, column=1).value          # the '#' column
        if not isinstance(num, (int, float)):
            if tops:
                break
            continue
        top = ws.cell(row=r, column=2).value          # Top Body (ft)
        if isinstance(top, (int, float)) and top >= 0:
            tops.append(top)
    return min(tops) if tops else None


def _shoe_hanger_mapping(pipe_model, excel_path):
    """Build {tag: text} for the shoe/hanger callouts from the pipe model.

    Reuses each pipe's pre-computed `name` (e.g. '18 5/8" Casing') and `shoe`;
    computes the liner hanger depth itself from the workbook. The deepest shoe
    (at/below the logged TD) is omitted."""
    mapping = {}
    shoes = [p["shoe"] for p in pipe_model
             if isinstance(p.get("shoe"), (int, float))]
    logged_td = max(shoes) if shoes else None

    wb = None
    for p in pipe_model:
        role = p.get("role")
        if not role:
            continue
        name = (p.get("name") or "").replace('"', _INCH)
        shoe = p.get("shoe")

        # Shoe callout — every pipe, except the bottom string (at/below TD).
        if isinstance(shoe, (int, float)):
            if logged_td is None or shoe < logged_td - _TD_TOL:
                mapping[f"{{{{ovl_shoe_{role}}}}}"] = (
                    f"{name} Shoe at {_format_depth(shoe)}ft"
                )

        # Hanger callout — liners only; hanger depth = top of the liner. Prefer
        # the XML top depth (authoritative, set on the pipe model); fall back to
        # the Excel min Top Body only if the XML didn't supply it.
        if p.get("type") == "LNR":
            hang = p.get("hanger")
            if hang is None:
                if wb is None:
                    wb = _load_wb(excel_path)
                hang = _min_top_body(wb, p.get("sheet"))
            if isinstance(hang, (int, float)):
                mapping[f"{{{{ovl_hanger_{role}}}}}"] = (
                    f"{name} Hanger at {_format_depth(hang)}ft"
                )
    return mapping


def _replace_tags_in_paragraph(p, mapping):
    """Replace mapping tags in one text-box paragraph, run-preserving on the
    first run (joining runs first so a split tag still matches). Returns 1 if the
    paragraph changed, else 0."""
    ts = p.findall(".//" + _W_T)
    if not ts:
        return 0
    joined = "".join(t.text or "" for t in ts)
    new = joined
    for tag, value in mapping.items():
        if tag in new:
            new = new.replace(tag, value)
    if new == joined:
        return 0
    ts[0].text = new
    ts[0].set(qn("xml:space"), "preserve")
    for t in ts[1:]:
        t.text = ""
    return 1


def _replace_in_textboxes(doc, mapping):
    """Apply `mapping` to every text box in the body (both the DrawingML copy and
    the VML fallback, so they stay in sync). Returns the count changed."""
    if not mapping:
        return 0
    changed = 0
    for txbx in doc.element.body.iter(_W_TXBX):
        for p in txbx.iter(_W_P):
            changed += _replace_tags_in_paragraph(p, mapping)
    return changed


def _remove_unfilled_boxes(doc):
    """Remove any text box still holding an unfilled {{ovl_shoe_*}} / {{ovl_hanger_*}}
    tag (absent pipe, omitted bottom shoe, non-liner hanger). Removes the whole
    enclosing run — which carries both the DrawingML and VML copies — and the
    paragraph if that leaves it empty. Returns the count removed."""
    body = doc.element.body
    runs, seen = [], set()
    for txbx in body.iter(_W_TXBX):
        text = "".join(t.text or "" for t in txbx.iter(_W_T))
        if not _REMOVABLE.search(text):
            continue
        run = txbx
        while run is not None and run.tag != _W_R:
            run = run.getparent()
        if run is None or id(run) in seen:
            continue
        seen.add(id(run))
        runs.append(run)

    removed = 0
    for run in runs:
        para = run.getparent()
        if para is None:
            continue
        para.remove(run)
        removed += 1
        if para.tag == _W_P:
            leftover = [c for c in para if c.tag != _W_PPR]
            if not leftover:
                gp = para.getparent()
                if gp is not None:
                    gp.remove(para)
    return removed


def _damage_mapping(clusters):
    """{tag: text} for the per-damage-point metal-loss and channel overlays.

    Block i (1-based, depth-sorted), point k (1-based within the block):
        {{ovl_ml<i>_<k>}} -> "<severity> metal loss in <suffix> Max WL% is
                              <loss>% at <depth>ft"
        {{ovl_ch<i>_<k>}} -> "Channel <n> is used to calculate <the ml text>."
    The channel overlay is emitted only when a channel was resolved."""
    mapping = {}
    for i, cluster in enumerate(clusters, start=1):
        for k, p in enumerate(cluster, start=1):
            suffix = (p.get("suffix") or "").replace('"', _INCH)
            ml = (f"{p.get('severity', '')} metal loss in {suffix} "
                  f"Max WL% is {float(p['loss']):.1f}% at {float(p['depth']):.1f}ft")
            mapping[f"{{{{ovl_ml{i}_{k}}}}}"] = ml
            ch = p.get("channel")
            if ch:
                mapping[f"{{{{ovl_ch{i}_{k}}}}}"] = (
                    f"Channel {ch} is used to calculate {ml}."
                )
    return mapping


# --- Public entry point ------------------------------------------------------
def apply_overlays(path, wellhead_damage=None, pipe_model=None, excel_path=None,
                   damage_clusters=None, progress=None, review=None, doc=None):
    """Fill overlay text boxes in the document.

    `wellhead_damage`: True -> damage statement, False -> clean, None -> skip.
    `pipe_model` + `excel_path`: enable the shoe/hanger callouts.
    `damage_clusters`: the damage-picture clusters — enables the per-point
    metal-loss / channel callouts inside each damage block.
    Unfilled variable slots (shoe/hanger/ml/ch) are removed. Returns boxes filled.

    When `doc` is given, operate on that live document and do not save (the
    caller owns the single open/save); otherwise open and save `path`. Taking a
    document object in is the module's only touch-point with the engine — no
    behaviour is shared — so it stays self-contained.
    """
    log = progress or (lambda m: None)
    rev = review or (lambda m: None)

    if wellhead_damage is None and not pipe_model and not damage_clusters:
        return 0

    mapping = {}
    if wellhead_damage is not None:
        mapping[WELLHEAD_TAG] = WELLHEAD_DAMAGE if wellhead_damage else WELLHEAD_CLEAN
    if pipe_model:
        mapping.update(_shoe_hanger_mapping(pipe_model, excel_path))
    if damage_clusters:
        mapping.update(_damage_mapping(damage_clusters))

    own = doc is None
    if own:
        doc = Document(path)
    filled = _replace_in_textboxes(doc, mapping)
    removed = _remove_unfilled_boxes(doc) if (pipe_model or damage_clusters) else 0
    if own and (filled or removed):
        doc.save(path)

    log(f"Overlays: filled {filled} box(es), removed {removed} unused.")
    if filled or removed:
        rev(f"Overlays — {filled} filled, {removed} removed.")
    return filled

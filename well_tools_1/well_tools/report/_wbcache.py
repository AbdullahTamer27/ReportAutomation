"""A tiny, transparent workbook cache for the report engine.

One report generation opens the same Excel workbook from several passes
(``tables``, ``charts``, ``pipe_config``, ``damage_select`` …). Each
``openpyxl.load_workbook`` re-parses the whole file, which for a formula-heavy
``.xlsm`` is seconds of redundant work.

``load()`` is a drop-in for ``openpyxl.load_workbook(path, data_only=True)`` that
returns a *shared, read-only* workbook and reloads only when the file changes on
disk (mtime or size). Consumers must treat the returned workbook as read-only —
they already do; every current caller only reads cell values. The isolated
``overlays`` module deliberately does NOT use this (it keeps its own load).

Keyed by ``(abspath, data_only, read_only)`` so distinct signatures don't clash;
a changed file transparently invalidates its entry, so a cached read is never
stale.
"""

import os

import openpyxl

# (abspath, data_only, read_only) -> (mtime, size, workbook)
_CACHE = {}


def load(path, data_only=True, read_only=False):
    """Return a cached workbook for `path`, reloading if the file changed.

    Behaves exactly like ``openpyxl.load_workbook(path, data_only=data_only,
    read_only=read_only)`` on a cold cache; on a warm cache it returns the same
    object without touching disk (beyond a cheap ``stat``)."""
    key = (os.path.abspath(path), bool(data_only), bool(read_only))
    st = os.stat(path)
    stamp = (st.st_mtime_ns, st.st_size)

    hit = _CACHE.get(key)
    if hit is not None and hit[0] == stamp[0] and hit[1] == stamp[1]:
        return hit[2]

    wb = openpyxl.load_workbook(path, data_only=data_only, read_only=read_only)
    _CACHE[key] = (stamp[0], stamp[1], wb)
    return wb


def clear():
    """Drop all cached workbooks (frees memory; next load re-reads from disk)."""
    _CACHE.clear()

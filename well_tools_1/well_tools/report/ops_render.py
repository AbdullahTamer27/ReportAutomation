"""Draw the one-page summary from the Excel template — without Excel.

The template is the *design*: a hand-formatted sheet whose column widths, row
heights, fonts, fills, borders and merges say exactly how the summary should
look. This module reads that geometry with openpyxl and redraws it with
PyMuPDF, substituting the well's values on the way.

Why not fill the workbook and have Excel render it, which is the obvious route:
because it makes Excel a hard dependency of every report, it cannot run on CI,
and it drags in a long tail of Excel-specific failures — macro-enabled content
types, merged ranges that don't move with an inserted row, hidden rows, COM
dialogs. None of that exists here. The template is only ever *read*.

The payoff shows up in the awkward part. Growing a table to fit a well with
twelve pipe sections is a list insertion in an in-memory model, not a
structural edit to a spreadsheet — so the rows below simply move, and there is
nothing to repair.
"""

import os
import re

from openpyxl.utils import get_column_letter

_TAG = re.compile(r"\{\{(\w+)\}\}")

# Excel's column width is in "characters" — about 7px each at 96dpi plus 5px of
# padding. Row heights are already in points, which is what a PDF wants.
_PX_PER_CHAR = 7
_CELL_PADDING_PX = 5
_PX_TO_PT = 72 / 96
_DEFAULT_COL_WIDTH = 8.43
_DEFAULT_ROW_HEIGHT = 15.0

# The base-14 fonts: built into every PDF reader and into PyMuPDF, so there is
# no font file to bundle and none to go missing from a frozen build. They stand
# in for the template's Calibri and Univers and run a little wider, which the
# shrink-to-fit in `_draw_cell` absorbs.
_FONT = "helv"
_FONT_BOLD = "hebo"

_LEFT, _CENTRE, _RIGHT = 0, 1, 2
_ALIGN = {"left": _LEFT, "center": _CENTRE, "centre": _CENTRE, "right": _RIGHT}

_TEXT_PAD_PT = 2.5
# White space between the panel and the log. Not a template column: the two
# halves are composed here, so the gap between them belongs here too.
_GUTTER_PT = 14
# Bullets are drawn, not typed: the base-14 fonts have no glyph for these.
_BULLET_CHARS = "•●"
# insert_textbox draws nothing at all if its box is a hair too short, so the
# space a line needs is measured generously.
_FIRST_LINE = 1.5
_NEXT_LINE = 1.25


class OpsRenderError(Exception):
    """The summary could not be drawn."""


# --------------------------------------------------------------------------
# Values
# --------------------------------------------------------------------------
# Stated on every well, whatever the data says.
TEMPERATURE_NOTE = "Temperature anomaly observed across the logging interval."


def _thickness(value):
    """Nominal thickness to three decimals — 0.250, not 0.25. The trailing zero
    is significant to a reader comparing it against a measured thickness."""
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)


def _number(value):
    """Whole numbers lose the decimal point (32.0 → 32); anything non-numeric —
    "Didn't Detect" is a legitimate cell value — passes through untouched."""
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):g}"
    except (TypeError, ValueError):
        return str(value)


def _od_value(label):
    """The numeric OD from a label like ``18 5/8" CSG`` → 18.625."""
    match = re.match(r'\s*(\d+)(?:\s+(\d+)\s*/\s*(\d+))?', str(label))
    if not match:
        return 0.0
    whole = float(match.group(1))
    if match.group(2):
        whole += float(match.group(2)) / float(match.group(3))
    return whole


def _is_tubing(label):
    return "TBG" in str(label).upper()


def sort_strings(rows):
    """Casings and liners together by OD descending, tubing always last.

    Tubing is the *innermost* string regardless of its diameter, so a 3 1/2"
    liner still belongs above a 4 1/2" tubing — which sorting by OD alone would
    get backwards."""
    return sorted(rows, key=lambda r: (_is_tubing(r.get("Pipe OD")),
                                       -_od_value(r.get("Pipe OD"))))


def sort_hotspots(hotspots):
    """The hot spots in the same order as the completion strings.

    They arrive in the pipe model's order, which follows the *configuration
    string* — and that is conventionally written inner→outer, so leaving it
    alone puts the tubing first and the biggest casing last. The summary reads
    largest-first with the tubing last, like every other table in the report."""
    return sorted(hotspots,
                  key=lambda h: (_is_tubing(h.get("pipe", {}).get("suffix")),
                                 -_od_value(h.get("pipe", {}).get("suffix"))))


def conclusion(pipe, grade):
    """One conclusions line, e.g. "Moderate metal loss detected across the
    7 5/8" liner string." — using the severity word the rest of the report uses,
    so the summary cannot describe a pipe differently from the tables."""
    from .pipe_config import SEVERITY, TYPE_FULL

    severity = SEVERITY.get(grade)
    if not severity:
        return None
    label = str(pipe.get("suffix", "")).rsplit(" ", 1)[0]   # size, minus the type code
    kind = TYPE_FULL.get(pipe.get("type"), "").lower()
    return f"{severity} metal loss detected across the {label} {kind} string."


# --------------------------------------------------------------------------
# The model
# --------------------------------------------------------------------------
class Cell:
    """One drawn cell: what it says and how the template says to draw it."""

    __slots__ = ("value", "size", "bold", "colour", "fill", "borders",
                 "align", "valign", "wrap", "colspan", "rowspan")

    def __init__(self, value=None, size=11.0, bold=False, colour=(0, 0, 0),
                 fill=None, borders=(), align=_LEFT, valign="center",
                 wrap=False, colspan=1, rowspan=1):
        self.value = value
        self.size = size
        self.bold = bold
        self.colour = colour
        self.fill = fill
        self.borders = set(borders)
        self.align = align
        self.valign = valign
        self.wrap = wrap
        self.colspan = colspan
        self.rowspan = rowspan

    def copy(self, value=None):
        clone = Cell(value, self.size, self.bold, self.colour, self.fill,
                     self.borders, self.align, self.valign, self.wrap,
                     self.colspan, self.rowspan)
        return clone


class Row:
    """A drawn row: its height and the cells across it, keyed by column index."""

    __slots__ = ("height", "cells")

    def __init__(self, height, cells=None):
        self.height = height
        self.cells = cells if cells is not None else {}

    def copy(self):
        return Row(self.height, {c: cell.copy(cell.value)
                                 for c, cell in self.cells.items()})

    def tags(self):
        found = []
        for cell in self.cells.values():
            if isinstance(cell.value, str):
                found.extend(_TAG.findall(cell.value))
        return found


class Layout:
    """The sheet as a grid of drawn rows, plus where the log image sits."""

    def __init__(self, columns, rows, image=None):
        self.columns = columns          # {column index: width in points}
        self.rows = rows                # list of Row, top to bottom
        self.image = image              # {"first_col","last_col","first_row","last_row"}

    @property
    def width(self):
        return sum(self.columns.values())

    @property
    def height(self):
        return sum(r.height for r in self.rows)

    def column_x(self, column):
        """The left edge of `column`, in points from the layout's left edge."""
        x = 0.0
        for index in sorted(self.columns):
            if index == column:
                return x
            x += self.columns[index]
        return x

    def find_tag(self, name):
        """``(row index, column index)`` of the first cell holding `name`."""
        for index, row in enumerate(self.rows):
            for column, cell in row.cells.items():
                if isinstance(cell.value, str) and name in _TAG.findall(cell.value):
                    return index, column
        return None


# --------------------------------------------------------------------------
# Reading the template
# --------------------------------------------------------------------------
def _rgb(colour):
    value = getattr(colour, "rgb", None)
    if not isinstance(value, str) or len(value) < 6:
        return None
    value = value[-6:]
    return tuple(int(value[i:i + 2], 16) / 255 for i in (0, 2, 4))


def _column_points(ws, column):
    dim = ws.column_dimensions.get(get_column_letter(column))
    width = dim.width if dim is not None and dim.width else _DEFAULT_COL_WIDTH
    return (width * _PX_PER_CHAR + _CELL_PADDING_PX) * _PX_TO_PT


def _print_area(ws):
    """The rectangle the template asks to be printed, as (r0, c0, r1, c1)."""
    area = ws.print_area
    if isinstance(area, (list, tuple)):
        area = area[0] if area else None
    if area:
        ref = str(area).replace("$", "").split("!")[-1]
        if ":" in ref:
            first, last = ref.split(":")
            return ws[first].row, ws[first].column, ws[last].row, ws[last].column
    # No print area set: fall back to everything that has content — including
    # the log image, whose columns hold no text and would otherwise be left out,
    # taking the entire right-hand half of the sheet with them.
    rows = [c.row for row in ws.iter_rows() for c in row if c.value not in (None, "")]
    cols = [c.column for row in ws.iter_rows() for c in row if c.value not in (None, "")]
    if not rows:
        raise OpsRenderError("the OPS template is empty")
    r0, c0, r1, c1 = min(rows), min(cols), max(rows), max(cols)
    for image in ws._images:
        anchor = image.anchor
        c0 = min(c0, anchor._from.col + 1)
        c1 = max(c1, anchor.to.col + 1)
        r0 = min(r0, anchor._from.row + 1)
    return r0, c0, r1, c1


def read_layout(path):
    """Read the template into a `Layout`.

    Hidden rows are dropped — the template keeps a few rows of sample data
    hidden, and Excel would not print them either."""
    from openpyxl import load_workbook

    if not os.path.isfile(path):
        raise OpsRenderError(f"OPS template not found: {path}")
    workbook = load_workbook(path)
    ws = workbook[workbook.sheetnames[0]]

    r0, c0, r1, c1 = _print_area(ws)
    columns = {c: _column_points(ws, c) for c in range(c0, c1 + 1)}

    spans = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        spans[(rng.min_row, rng.min_col)] = (rng.max_row - rng.min_row + 1,
                                             rng.max_col - rng.min_col + 1)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))

    rows = []
    for r in range(r0, r1 + 1):
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.hidden:
            continue
        height = dim.height if dim is not None and dim.height else _DEFAULT_ROW_HEIGHT
        cells = {}
        for c in range(c0, c1 + 1):
            if (r, c) in covered:
                continue
            source = ws.cell(row=r, column=c)
            font, fill, border, align = (source.font, source.fill,
                                         source.border, source.alignment)
            sides = {side for side in ("left", "right", "top", "bottom")
                     if getattr(border, side) is not None
                     and getattr(border, side).style}
            rowspan, colspan = spans.get((r, c), (1, 1))
            value = source.value
            if isinstance(value, str):
                # A trailing newline is an authoring artifact — Excel shows it as
                # nothing, but here it counts as a second line and pushes the
                # text off its own vertical alignment. Leading space is kept:
                # some labels are indented deliberately.
                value = value.rstrip()
            cells[c] = Cell(
                value=value,
                size=float(font.sz) if font.sz else 11.0,
                bold=bool(font.b),
                colour=_rgb(font.color) or (0, 0, 0) if font.color else (0, 0, 0),
                fill=_rgb(fill.fgColor) if fill.fill_type == "solid" else None,
                borders=sides,
                align=_ALIGN.get((align.horizontal or "").lower(), _LEFT),
                valign=align.vertical or "center",
                wrap=bool(align.wrap_text),
                colspan=colspan, rowspan=rowspan,
            )
        rows.append(Row(height, cells))

    image = None
    if ws._images:
        anchor = ws._images[0].anchor
        image = {"first_col": anchor._from.col + 1, "last_col": anchor.to.col + 1,
                 "first_row": anchor._from.row + 1, "last_row": anchor.to.row + 1}
    return Layout(columns, rows, image)


# --------------------------------------------------------------------------
# Filling
# --------------------------------------------------------------------------
def _substitute(text, values, defaults):
    """Replace the tags in one cell's text. Returns ``(text, any value found)``;
    a cell whose tags are all empty is dropped, so a well with no workover shows
    no workover line rather than "Latest Workover Date: "."""
    found = False
    for name in _TAG.findall(text):
        value = str(values.get(name.lower(), "") or "").strip()
        if value.upper() == "N/A":
            value = ""
        if not value:
            value = str(defaults.get(name.lower(), "") or "")
        if value:
            found = True
        text = text.replace("{{%s}}" % name, value)
    return text, found


def _fill_scalars(layout, fields, defaults):
    """Substitute the single-value tags, dropping rows that come back empty."""
    values = {str(k).lower(): v for k, v in fields.items()}
    defaults = {str(k).lower(): v for k, v in (defaults or {}).items()}
    block_tags = set(STRINGS_TAGS) | set(HOTSPOT_TAGS) | {CONCLUSION_TAG}

    kept = []
    for row in layout.rows:
        tags = [t for t in row.tags()]
        if tags and not any(t in block_tags for t in tags):
            any_value = False
            for cell in row.cells.values():
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value, found = _substitute(cell.value, values, defaults)
                    any_value = any_value or found
            if not any_value:
                continue                      # nothing to say — drop the row
        kept.append(row)
    layout.rows = kept


def _widen_for_content(layout, column, rows, limit=1.5):
    """Widen `column` so the longest label in `rows` fits on one line.

    The template fixes the column widths, but it cannot know what a well will be
    called: a tapered string's name — ``4 1/2" × 3 1/2" × 2 7/8" TBG`` — is far
    longer than a plain ``7" CSG`` and wraps to two cramped lines. Only the panel
    grows; the log's width follows its own height, so nothing is squeezed to pay
    for it. Capped at `limit`× so one absurd label cannot distort the page."""
    import fitz

    current = layout.columns.get(column)
    if not current:
        return
    needed = current
    for index in rows:
        cell = layout.rows[index].cells.get(column)
        if cell is None or cell.value in (None, ""):
            continue
        font = fitz.Font(_FONT_BOLD if cell.bold else _FONT)
        width = font.text_length(str(cell.value), cell.size) + 2 * _TEXT_PAD_PT
        needed = max(needed, width)
    layout.columns[column] = min(needed, current * limit)


def _merge_repeated(layout, first, count, column):
    """Merge `column` down across consecutive rows holding the same value.

    A pipe in three weight sections occupies three rows; repeating its OD in
    each reads as three separate pipes. In the layout a merge is just a rowspan
    on the first cell and dropping the ones it covers — the spanned cell's
    borders are then drawn around the whole group, which is the look Excel gives
    a merged range."""
    start = 0
    while start < count:
        head = layout.rows[first + start].cells.get(column)
        end = start
        while end + 1 < count:
            following = layout.rows[first + end + 1].cells.get(column)
            if head is None or following is None or following.value != head.value:
                break
            end += 1
        if end > start and head is not None:
            head.rowspan = end - start + 1
            for offset in range(start + 1, end + 1):
                layout.rows[first + offset].cells.pop(column, None)
        start = end + 1


def _grow_block(layout, first_tag, columns_by_tag, records, formatter):
    """Replace the tagged row with one row per record.

    This is the whole reason for rendering rather than writing a workbook: the
    table grows by building a list, so the rows below just move down and there
    is no spreadsheet structure to keep consistent."""
    found = layout.find_tag(first_tag)
    if found is None:
        return 0, None
    index, _ = found
    template_row = layout.rows[index]

    # The blank rows under the tag are absorbed. They carry the table's own
    # borders and fills, so leaving them in draws empty table rows under the
    # data; spacing before the next heading comes from that heading's own row
    # height, which is a thing the template can say.
    end = index + 1
    while end < len(layout.rows) and not any(
            c.value not in (None, "") for c in layout.rows[end].cells.values()):
        end += 1

    built = []
    for record in records:
        row = template_row.copy()
        for tag, column in columns_by_tag.items():
            if column in row.cells:
                row.cells[column].value = formatter(record, tag)
        built.append(row)

    layout.rows[index:end] = built
    return len(built), index


# The tags that make up one row of each repeating block, in column order.
STRINGS_TAGS = ("str_od", "str_weight", "str_top", "str_bottom", "str_nom")
HOTSPOT_TAGS = ("hs_od", "hs_wl", "hs_grade", "hs_depth")
CONCLUSION_TAG = "conclusions"


# --------------------------------------------------------------------------
# Drawing
# --------------------------------------------------------------------------
def _line_height(lines, size):
    return size * (_FIRST_LINE + max(0, lines - 1) * _NEXT_LINE)


def _wrapped_lines(text, font, size, width):
    """How many lines `text` takes at `size` in `width` — a greedy word wrap,
    matching how the text box will break it."""
    if width <= 0:
        return 1
    total = 0
    for paragraph in str(text).split("\n"):
        line, count = "", 1
        for word in paragraph.split():
            candidate = f"{line} {word}".strip()
            if line and font.text_length(candidate, size) > width:
                line, count = word, count + 1
            else:
                line = candidate
        total += count
    return max(1, total)


def _draw_cell(page, rect, cell):
    import fitz

    if cell.fill is not None:
        page.draw_rect(rect, color=None, fill=cell.fill)
    for side in cell.borders:
        if side == "left":
            a, b = (rect.x0, rect.y0), (rect.x0, rect.y1)
        elif side == "right":
            a, b = (rect.x1, rect.y0), (rect.x1, rect.y1)
        elif side == "top":
            a, b = (rect.x0, rect.y0), (rect.x1, rect.y0)
        else:
            a, b = (rect.x0, rect.y1), (rect.x1, rect.y1)
        page.draw_line(fitz.Point(*a), fitz.Point(*b), color=(0, 0, 0), width=0.7)

    text = cell.value
    if text in (None, ""):
        return
    text = str(text)

    # The base-14 fonts have no bullet glyph and draw "?" instead, in any
    # encoding — so a leading bullet is drawn as a filled circle and the text
    # indented past it.
    left_pad = _TEXT_PAD_PT
    if text[:1] in _BULLET_CHARS:
        text = text[1:].lstrip()
        radius = cell.size * 0.13
        centre = fitz.Point(rect.x0 + _TEXT_PAD_PT + radius,
                            rect.y0 + rect.height / 2)
        page.draw_circle(centre, radius, color=(0, 0, 0), fill=(0, 0, 0))
        left_pad += 2 * radius + cell.size * 0.22

    inner_width = rect.width - left_pad - _TEXT_PAD_PT
    size = cell.size
    # Shrink to fit rather than spill. The template cannot know how long a well
    # name will be, and the substituted faces are wider than the authored ones,
    # so text that fitted in Excel can overflow here.
    font = fitz.Font(_FONT_BOLD if cell.bold else _FONT)
    if cell.wrap:
        # Wrapped text overflows *downwards*, and a cell clipped at the row
        # boundary loses a whole line — "Hot Spot Depth" showing as "Hot Spot".
        while size > 5 and _line_height(_wrapped_lines(text, font, size, inner_width),
                                        size) > rect.height:
            size -= 0.5
    else:
        while size > 5 and font.text_length(text, size) > inner_width:
            size -= 0.5

    lines = _wrapped_lines(text, font, size, inner_width) if cell.wrap \
        else text.count("\n") + 1
    needed = _line_height(lines, size)
    if cell.valign == "top":
        top = rect.y0 + _TEXT_PAD_PT
    elif cell.valign == "bottom":
        top = max(rect.y0, rect.y1 - needed)
    else:
        top = rect.y0 + max(0, (rect.height - needed) / 2)

    box = fitz.Rect(rect.x0 + left_pad, top,
                    rect.x1 - _TEXT_PAD_PT, rect.y1 + needed)
    page.insert_textbox(box, text, fontname=(_FONT_BOLD if cell.bold else _FONT),
                        fontsize=size, align=cell.align, color=cell.colour)


def draw(layout, dest_path, image_path=None, dpi=200):
    """Draw `layout` to `dest_path` as a PNG. Returns ``{"path", "size"}``."""
    import fitz

    height = layout.height
    if height <= 0:
        raise OpsRenderError("the OPS layout has no size")

    # The log sets the width of its own half. Its columns in the template are
    # only a placeholder box; drawn at the panel's full height and its own
    # proportions, the two halves finish level with each other whatever the
    # well's row count did to the panel — which is the thing that could not be
    # arranged from inside Excel, where the image is pinned to a cell range.
    panel_width = sum(w for c, w in layout.columns.items()
                      if layout.image is None or c < layout.image["first_col"])
    log_width = 0.0
    if image_path and layout.image and os.path.isfile(image_path):
        try:
            from PIL import Image as _PILImage

            with _PILImage.open(image_path) as probe:
                log_w, log_h = probe.size
            log_width = height * (log_w / log_h) if log_h else 0.0
        except Exception:  # noqa: BLE001 — a bad log image must not stop the page
            log_width = 0.0

    gutter = _GUTTER_PT if log_width > 0 else 0
    width = panel_width + gutter + log_width
    if width <= 0:
        raise OpsRenderError("the OPS layout has no size")

    doc = fitz.open()
    try:
        page = doc.new_page(width=width, height=height)
        page.draw_rect(fitz.Rect(0, 0, width, height), color=None, fill=(1, 1, 1))

        tops = []
        y = 0.0
        for row in layout.rows:
            tops.append(y)
            y += row.height

        for index, row in enumerate(layout.rows):
            for column, cell in row.cells.items():
                x0 = layout.column_x(column)
                x1 = x0 + sum(layout.columns.get(column + i, 0)
                              for i in range(cell.colspan))
                y0 = tops[index]
                last = min(index + cell.rowspan - 1, len(layout.rows) - 1)
                y1 = tops[last] + layout.rows[last].height
                _draw_cell(page, fitz.Rect(x0, y0, x1, y1), cell)

        if log_width > 0:
            page.insert_image(fitz.Rect(panel_width + gutter, 0, width, height),
                              filename=image_path, keep_proportion=True)

        zoom = dpi / 72.0
        pixmap = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        pixmap.save(dest_path)
        size = (pixmap.width, pixmap.height)
    finally:
        doc.close()
    return {"path": dest_path, "size": size}


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------
def render_ops(template_path, dest_path, fields, pipe_summary_rows, hotspots,
               proc_path=None, defaults=None, dpi=200):
    """Draw the whole one-page summary. Returns ``{"path", "size", "warnings"}``.

    The template is read, never written; nothing here needs Excel."""
    from .tables import GRADE_COLORS

    layout = read_layout(template_path)
    warnings = []

    _fill_scalars(layout, fields, defaults)

    strings = sort_strings(list(pipe_summary_rows))
    columns = {}
    for tag in STRINGS_TAGS:
        found = layout.find_tag(tag)
        if found:
            columns[tag] = found[1]
    keys = {"str_od": "Pipe OD", "str_weight": "Weight (ppf)", "str_top": "Top (ft)",
            "str_bottom": "Bottom (ft)", "str_nom": "Thick_Nom"}

    def string_value(record, tag):
        raw = record.get(keys[tag])
        if tag == "str_nom":
            return _thickness(raw)
        if tag == "str_od":
            return raw
        value = _number(raw)
        return f"{value:g}" if isinstance(value, float) else value

    count, first = _grow_block(layout, "str_od", columns, strings, string_value)
    if first is not None and "str_od" in columns:
        _widen_for_content(layout, columns["str_od"], range(first, first + count))
        _merge_repeated(layout, first, count, columns["str_od"])

    # Largest first, tubing last — the same order as the strings table above.
    # The conclusions are built from this list, so they follow it.
    hotspots = sort_hotspots(list(hotspots))

    hs_columns = {}
    for tag in HOTSPOT_TAGS:
        found = layout.find_tag(tag)
        if found:
            hs_columns[tag] = found[1]

    def hotspot_value(spot, tag):
        if tag == "hs_od":
            return spot["pipe"].get("suffix")
        if tag == "hs_wl":
            try:
                return f"{float(str(spot.get('max_loss')).rstrip('%')):.1f}%"
            except (TypeError, ValueError):
                return spot.get("max_loss")
        if tag == "hs_grade":
            return spot.get("grade")
        return spot.get("depth")

    grade_column = hs_columns.get("hs_grade")
    count, first = _grow_block(layout, "hs_od", hs_columns, hotspots, hotspot_value)
    if first is not None and "hs_od" in hs_columns:
        # A tapered string's combined name lands here, and it is the longest
        # label the engine can produce.
        _widen_for_content(layout, hs_columns["hs_od"], range(first, first + count))
    if first is not None and grade_column is not None:
        for offset in range(count):
            grade = hotspots[offset].get("grade")
            code = GRADE_COLORS.get(grade)
            cell = layout.rows[first + offset].cells.get(grade_column)
            if code and cell is not None:
                cell.fill = _rgb(type("C", (), {"rgb": code})())

    lines = [c for c in (conclusion(s["pipe"], s.get("grade")) for s in hotspots) if c]
    lines.append(TEMPERATURE_NOTE)
    _grow_block(layout, CONCLUSION_TAG,
                {CONCLUSION_TAG: (layout.find_tag(CONCLUSION_TAG) or (0, 0))[1]},
                lines, lambda line, tag: line)

    result = draw(layout, dest_path, image_path=proc_path, dpi=dpi)
    result["warnings"] = warnings
    return result

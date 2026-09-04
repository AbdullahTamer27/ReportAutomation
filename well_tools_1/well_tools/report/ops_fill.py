"""Fill the OPS workbook — the one-page summary, built in Excel.

The template (``webapp/data/ops/OPS.xlsx``) is a hand-formatted sheet carrying
``{{tags}}``. This module writes the well's values into a copy of it and leaves
the formatting alone, so the design lives in Excel where it can be seen, and the
code only supplies numbers.

Each repeating block — completion strings, hot spots, conclusions — is one
tagged row in the template that grows to however many rows the well needs. It
first uses any blank rows already sitting under the tag, then inserts more; any
left spare are hidden, and hidden rows don't print. So the template needs no
reserved space and no count kept in sync with this file.

Growing a sheet is only safe because of one detail, in `insert_rows`: merged
ranges below the insertion are **shifted in place**. Unmerging and remerging
them instead — the obvious repair — silently destroys cell values, because
``merge_cells()`` blanks everything but a range's top-left corner.

Values come from the same places the Word report uses — ``build_pipe_summary``
for the completion strings, ``tables.worst_joint`` for the hot spots, and
``pipe_config.SEVERITY`` for the conclusion wording — so the summary cannot
disagree with the report it summarises.
"""

import copy
import os
import re

from openpyxl.cell.cell import MergedCell

from .pipe_config import TYPE_FULL
from .tables import GRADE_COLORS
from .ops_panel import TEMPERATURE_NOTE, _conclusion

_TAG = re.compile(r"\{\{(\w+)\}\}")

# Which tag marks the first row of each repeating block, and the tags that make
# up one row of it, in column order.
STRINGS_TAGS = ("str_od", "str_weight", "str_top", "str_bottom", "str_nom")
HOTSPOT_TAGS = ("hs_od", "hs_wl", "hs_grade", "hs_depth")
CONCLUSION_TAG = "conclusions"

# Stray artwork: anything anchored beyond this column is outside the panel and
# the log, and is not part of the picture.
_PANEL_LAST_COL = 16


class OpsFillError(Exception):
    """The OPS workbook could not be filled."""


# --------------------------------------------------------------------------
# Sorting
# --------------------------------------------------------------------------
def _od_value(label):
    """The numeric OD from a label like ``18 5/8" CSG`` → 18.625."""
    match = re.match(r'\s*(\d+)(?:\s+(\d+)\s*/\s*(\d+))?', str(label))
    if not match:
        return 0.0
    whole = float(match.group(1))
    if match.group(2):
        whole += float(match.group(2)) / float(match.group(3))
    return whole


def _is_tubing(label):
    return "TBG" in str(label).upper()


def sort_strings(rows):
    """Casings and liners together by OD descending, tubing always last.

    Tubing is the *innermost* string regardless of its diameter, so a 3 1/2"
    liner still belongs above a 4 1/2" tubing — which sorting by OD alone would
    get backwards."""
    return sorted(rows, key=lambda r: (_is_tubing(r.get("Pipe OD")),
                                       -_od_value(r.get("Pipe OD"))))


# --------------------------------------------------------------------------
# Sheet helpers
# --------------------------------------------------------------------------
def _find_tags(ws):
    """``{tag name: (row, column)}`` for every ``{{tag}}`` in the sheet."""
    found = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                for name in _TAG.findall(cell.value):
                    found.setdefault(name, (cell.row, cell.column))
    return found


def _row_has_content(ws, row, first_col, last_col):
    """Read without instantiating. ``ws.cell()`` *creates* the cell it is asked
    for, which pushes ``max_row`` down — enough to make a scan that uses it as a
    bound run away forever."""
    for col in range(first_col, last_col + 1):
        cell = ws._cells.get((row, col))
        if cell is not None and cell.value not in (None, ""):
            return True
    return False


_UNBOUNDED = 500


def _capacity(ws, first_row, first_col, last_col):
    """How many rows this block may use: the tagged row plus the blank rows
    under it, stopping at the next row that holds anything.

    A block with nothing below it — Conclusions, usually — is unbounded, since
    there is no section underneath for it to run into. Rows past the formatted
    ones come out unstyled, which is better than dropping a conclusion."""
    limit = ws.max_row                      # snapshot: the sheet must not grow
    for row in range(first_row + 1, limit + 1):
        if _row_has_content(ws, row, first_col, last_col):
            return row - first_row
    return _UNBOUNDED


def insert_rows(ws, at, count, style_row):
    """Insert `count` rows before `at`, taking everything below down with them.

    openpyxl's own ``insert_rows`` moves cell *values* and nothing else: merged
    ranges, image anchors and row heights stay where they were, which tears a
    sheet apart. The repair matters in one specific way — the merged ranges are
    **shifted in place**, never unmerged and remerged. ``merge_cells()`` blanks
    every cell in the range but its top-left, so remerging a shifted range
    silently destroys the values that just moved into it.

    New rows inherit the tagged row's formatting, height, and any merge that
    lives entirely within it (the grade cell spans two columns)."""
    if count <= 0:
        return

    props = {r: (d.height, d.hidden) for r, d in ws.row_dimensions.items()}
    row_merges = [(rng.min_col, rng.max_col) for rng in ws.merged_cells.ranges
                  if rng.min_row == rng.max_row == style_row]
    anchors = [(img, img.anchor._from.row, getattr(img.anchor, "to", None))
               for img in ws._images]

    ws.insert_rows(at, count)

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= at:
            rng.shift(0, count)

    for row in sorted((r for r in props if r >= at), reverse=True):
        height, hidden = props[row]
        ws.row_dimensions[row + count].height = height
        ws.row_dimensions[row + count].hidden = hidden

    style_height = props.get(style_row, (None, False))[0]
    for offset in range(count):
        row = at + offset
        ws.row_dimensions[row].height = style_height
        ws.row_dimensions[row].hidden = False
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col)._style = copy.copy(
                ws.cell(row=style_row, column=col)._style)
        for min_col, max_col in row_merges:
            ws.merge_cells(start_row=row, start_column=min_col,
                           end_row=row, end_column=max_col)

    for img, frm, to in anchors:
        if frm >= at - 1:                       # anchors are 0-indexed
            img.anchor._from.row = frm + count
        if to is not None and to.row >= at - 1:
            to.row = to.row + count


def _set(ws, row, col, value):
    """Write a value, unhiding the row — a template row may start hidden."""
    ws.cell(row=row, column=col).value = value
    ws.row_dimensions[row].hidden = False


def _hide(ws, row):
    ws.row_dimensions[row].hidden = True


def _number(value):
    """Numbers as numbers, so the cell's own format governs how they display."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


# --------------------------------------------------------------------------
# Blocks
# --------------------------------------------------------------------------
def _fill_scalars(ws, tags, fields, defaults, warnings):
    """Substitute the single-value tags inside their sentences, and hide a row
    whose tags all come back empty — a well with no workover gets no line,
    rather than a line reading "Latest Workover Date: ."."""
    # Tags are written the way the template author wrote them ({{RIG}}), while
    # payload keys are always lowercase ("rig") — so match without case.
    values_by_key = {str(k).lower(): v for k, v in fields.items()}
    defaults_by_key = {str(k).lower(): v for k, v in defaults.items()}

    handled = set()
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or "{{" not in cell.value:
                continue
            names = _TAG.findall(cell.value)
            if not names or any(n in STRINGS_TAGS or n in HOTSPOT_TAGS
                                or n == CONCLUSION_TAG for n in names):
                continue
            text = cell.value
            values = []
            for name in names:
                value = str(values_by_key.get(name.lower(), "") or "").strip()
                if value.upper() == "N/A":
                    value = ""
                if not value:
                    # A field may answer for itself when left blank — no rig
                    # entered is stated as "Rigless" rather than left out. The
                    # wording belongs to the field registry, so it is passed in
                    # rather than repeated here.
                    value = str(defaults_by_key.get(name.lower(), "") or "")
                values.append(value)
                text = text.replace("{{%s}}" % name, value)
            if not any(values):
                cell.value = None
                _hide(ws, cell.row)
            else:
                cell.value = text
                ws.row_dimensions[cell.row].hidden = False
            handled.update(names)
    return handled


def _fill_table(ws, first_row, columns, values):
    """Write `values` (a list of per-row lists) down from `first_row`.

    The block grows to fit: it uses whatever blank rows already sit under the
    tagged row and inserts the rest, so the template only ever needs the one
    tagged row. Any spare rows left over are hidden, and hidden rows don't
    print.

    `columns` are the tags' own column numbers, not a contiguous span: the grade
    cell is merged across two columns, so the column beside it is a read-only
    MergedCell and the next value belongs one further along."""
    room = _capacity(ws, first_row, min(columns), max(columns))
    needed = len(values)

    if needed > room and room < _UNBOUNDED:
        insert_rows(ws, first_row + room, needed - room, first_row)
        room = needed

    # Give every row of the block the tagged row's own merges. Inserted rows get
    # them from `insert_rows`, but blank rows the template author left behind may
    # have been merged inconsistently — the grade cell spans two columns, and one
    # row missing that merge is visible in the finished picture.
    row_merges = [(rng.min_col, rng.max_col) for rng in ws.merged_cells.ranges
                  if rng.min_row == rng.max_row == first_row]
    existing = {(rng.min_row, rng.min_col) for rng in ws.merged_cells.ranges}
    for offset in range(needed):
        row = first_row + offset
        for min_col, max_col in row_merges:
            if (row, min_col) not in existing:
                ws.merge_cells(start_row=row, start_column=min_col,
                               end_row=row, end_column=max_col)

    for offset in range(needed):
        for col, value in zip(columns, values[offset]):
            _set(ws, first_row + offset, col, value)
    for offset in range(needed, min(room, _UNBOUNDED)):
        for col in columns:
            cell = ws.cell(row=first_row + offset, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        _hide(ws, first_row + offset)
    return needed


def _merge_repeated_od(ws, first_row, col, count, labels):
    """Merge the Pipe OD cell across the rows of one pipe.

    A pipe with three weight sections occupies three rows; leaving its OD
    repeated in each reads as three separate pipes."""
    start = 0
    while start < count:
        end = start
        while end + 1 < count and labels[end + 1] == labels[start]:
            end += 1
        if end > start:
            ws.merge_cells(start_row=first_row + start, start_column=col,
                           end_row=first_row + end, end_column=col)
        start = end + 1


# --------------------------------------------------------------------------
# Pictures
# --------------------------------------------------------------------------
def _place_proc(ws, proc_path, last_row):
    """Swap in this well's processed-log image, keeping the template's anchor,
    and pull its bottom edge to `last_row` so it ends level with the panel."""
    from openpyxl.drawing.image import Image as XLImage

    panel = [i for i, img in enumerate(ws._images)
             if img.anchor._from.col < _PANEL_LAST_COL]
    if not panel:
        return False
    index = panel[0]
    anchor = ws._images[index].anchor
    anchor.to.row = last_row - 1                  # anchors are 0-indexed

    image = XLImage(proc_path)
    image.anchor = anchor
    ws._images[index] = image
    return True


def _drop_stray_images(ws):
    """Remove artwork anchored outside the panel and the log."""
    before = len(ws._images)
    ws._images = [img for img in ws._images
                  if img.anchor._from.col < _PANEL_LAST_COL]
    return before - len(ws._images)


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------
def fill_ops(template_path, dest_path, fields, pipe_summary_rows, hotspots,
             proc_path=None, defaults=None):
    """Write a filled OPS workbook to `dest_path`. Returns
    ``{"path", "warnings"}``; the template is never modified.

    `defaults` maps a tag name to what it should say when the user leaves that
    field blank (e.g. ``{"RIG": "Rigless"}``). It comes from the field registry,
    which the engine cannot import, so the caller supplies it."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    if not os.path.isfile(template_path):
        raise OpsFillError(f"OPS template not found: {template_path}")

    warnings = []
    workbook = load_workbook(template_path)
    ws = workbook[workbook.sheetnames[0]]
    tags = _find_tags(ws)

    for required in ("str_od", "hs_od", CONCLUSION_TAG):
        if required not in tags:
            raise OpsFillError(f"the OPS template has no {{{{{required}}}}} tag")

    _fill_scalars(ws, tags, fields, defaults or {}, warnings)

    # --- completion strings -------------------------------------------------
    rows = sort_strings(list(pipe_summary_rows))
    first_row, first_col = tags["str_od"]
    str_cols = [tags[name][1] for name in STRINGS_TAGS if name in tags]
    values = [[r.get("Pipe OD"), _number(r.get("Weight (ppf)")),
               _number(r.get("Top (ft)")), _number(r.get("Bottom (ft)")),
               _number(r.get("Thick_Nom"))] for r in rows]
    used = _fill_table(ws, first_row, str_cols, values)
    _merge_repeated_od(ws, first_row, first_col, used,
                       [r.get("Pipe OD") for r in rows])

    # --- hot spots ----------------------------------------------------------
    # A block that grew pushed everything below it down, so the positions found
    # before are stale. Re-read them rather than trying to track the offset.
    tags = _find_tags(ws)
    hs_row, hs_col = tags["hs_od"]
    hs_cols = [tags[name][1] for name in HOTSPOT_TAGS if name in tags]
    hs_values = []
    for spot in hotspots:
        loss = spot.get("max_loss")
        # The template's Max WL cell is a percentage format, which expects a
        # fraction; the engine carries the value in percent units.
        try:
            loss = float(str(loss).rstrip("%")) / 100.0
        except (TypeError, ValueError):
            pass
        hs_values.append([spot["pipe"].get("suffix"), loss, spot.get("grade"),
                          _number(spot.get("depth"))])
    hs_used = _fill_table(ws, hs_row, hs_cols, hs_values)

    # grade colours, from the palette the Word tables use
    grade_col = tags["hs_grade"][1] if "hs_grade" in tags else hs_col + 2
    for offset in range(hs_used):
        grade = hotspots[offset].get("grade")
        code = GRADE_COLORS.get(grade)
        if code:
            ws.cell(row=hs_row + offset, column=grade_col).fill = PatternFill(
                "solid", fgColor=code)

    # --- conclusions --------------------------------------------------------
    lines = [c for c in (_conclusion(s["pipe"], s.get("grade")) for s in hotspots) if c]
    lines.append(TEMPERATURE_NOTE)
    tags = _find_tags(ws)                     # the hot-spot block may have grown
    c_row, c_col = tags[CONCLUSION_TAG]
    _fill_table(ws, c_row, [c_col], [[line] for line in lines])

    # --- pictures and page --------------------------------------------------
    dropped = _drop_stray_images(ws)
    last_row = max(r for r in range(1, ws.max_row + 1)
                   if not ws.row_dimensions[r].hidden
                   and _row_has_content(ws, r, 1, 8))
    if proc_path and os.path.isfile(proc_path):
        if not _place_proc(ws, proc_path, last_row):
            warnings.append("OPS: no log image placeholder found in the template")
    elif proc_path:
        warnings.append(f"OPS: log image not found — {os.path.basename(proc_path)}")

    right = get_column_letter(_PANEL_LAST_COL)
    ws.print_area = f"B4:{right}{last_row}"

    workbook.save(dest_path)
    return {"path": dest_path, "warnings": warnings, "stray_images": dropped}

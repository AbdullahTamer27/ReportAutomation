"""Interval Generator service.

Mirrors the desktop Interval Generator's "update template in place" path, reusing
the same `well_tools.core` functions (unmodified): parse a WellSchematic XML,
build the interval + pipe-summary tables (with optional THICKNESS-based
Channel/Offset rows), and write/overwrite the 'Raw Data' sheet in the given Excel
template. The rest of the workbook is untouched.
"""

import os
import logging

logger = logging.getLogger("webapp.interval")


class IntervalInputError(Exception):
    """Raised when the XML or template inputs are missing or invalid."""


def _build_preview_text(interval_df, pipe_summary_df):
    """Recreate the desktop tool's text preview of the generated tables."""
    lines = ["===== TABLE 1 — INTERVALS =====", ""]
    for i, row in interval_df.iterrows():
        lines.append(f"Interval {i + 1}: {row['Start Depth (ft)']} – {row['End Depth (ft)']} ft")
        for cfg in row["Configurations"]:
            lines.append(f"   • {cfg}")
        if "Channels" in interval_df.columns:
            lines.append(f"   Channel: {'-'.join(str(v) for v in row['Channels'])}")
            lines.append(f"   Offset:  {'/'.join(str(v) for v in row['Offsets'])}")
        lines.append("")
    lines.append("===== TABLE 2 — PIPE SUMMARY =====")
    lines.append("")
    lines.append(pipe_summary_df.to_string(index=False))
    return "\n".join(lines)


def generate_raw_data(xml_path, template_path):
    """Parse the XML and update the 'Raw Data' sheet of the template in place.

    Returns a summary dict. Raises IntervalInputError for bad inputs and lets
    PermissionError propagate (file open in Excel).
    """
    from well_tools.core.xml_parser import parse_wellschematic_xml, build_pipe_summary
    from well_tools.core.intervals import build_intervals_from_xml
    from well_tools.core.thickness import parse_thickness_sections
    from well_tools.core.excel_output import write_raw_data_to_template

    if not xml_path or not os.path.isfile(xml_path):
        raise IntervalInputError("XML file not found. Please choose a WellSchematic .xml file.")
    if not xml_path.lower().endswith(".xml"):
        raise IntervalInputError("The schematic must be a .xml file.")
    if not template_path or not os.path.isfile(template_path):
        raise IntervalInputError("Excel template not found. Please choose a .xlsx/.xlsm template.")
    if not template_path.lower().endswith((".xlsx", ".xlsm")):
        raise IntervalInputError("The template must be a .xlsx or .xlsm file.")

    xml_data = parse_wellschematic_xml(xml_path)

    # Optional THICKNESS sheet → Channel/Offset rows (same handling as desktop).
    thickness_sections = None
    try:
        sections = parse_thickness_sections(template_path)
        if sections:
            thickness_sections = sections
            thickness_note = (f"THICKNESS sheet found ({len(sections)} sections) "
                              "— Channel/Offset rows added.")
        else:
            thickness_note = ("THICKNESS sheet present but unreadable "
                              "— Channel/Offset rows omitted.")
    except ValueError as te:
        if str(te) == "NO_THICKNESS_SHEET":
            thickness_note = "No THICKNESS sheet — Channel/Offset rows omitted."
        else:
            thickness_note = f"THICKNESS read issue: {te} — rows omitted."

    pipe_summary_df = build_pipe_summary(xml_data)
    interval_df = build_intervals_from_xml(xml_data, thickness_sections=thickness_sections)

    write_raw_data_to_template(template_path, pipe_summary_df, interval_df)

    pipe_types = {str(k): int(v) for k, v in xml_data["Type"].value_counts().to_dict().items()}
    logger.info("Interval: %d pipes, %d intervals -> %s",
                len(xml_data), len(interval_df), template_path)

    return {
        "template_path": template_path,
        "num_pipes": int(len(xml_data)),
        "pipe_types": pipe_types,
        "depth_min": float(xml_data["Start"].min()),
        "depth_max": float(xml_data["End"].max()),
        "num_intervals": int(len(interval_df)),
        "thickness_note": thickness_note,
        "preview": _build_preview_text(interval_df, pipe_summary_df),
    }


INTERVALS_MAIN_SHEET = "intervals MAIN"


def _copy_sheet_snapshot(src_ws, dst_wb, title):
    """Copy a worksheet's values + common formatting into `dst_wb` as `title`.
    A static snapshot (values, not formulas) so the standalone file has no broken
    cross-sheet references. Read-only on the source."""
    from copy import copy
    dst = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.fill = copy(cell.fill)
                nc.border = copy(cell.border)
                nc.alignment = copy(cell.alignment)
                nc.number_format = cell.number_format
    for col, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst.column_dimensions[col].width = dim.width
    for r, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[r].height = dim.height
    for merged in src_ws.merged_cells.ranges:
        dst.merge_cells(str(merged))
    return dst


def generate_raw_data_file(xml_path, output_path, data_excel=None):
    """Build the Raw Data table from the XML and write it to a BRAND-NEW Excel at
    `output_path`, alongside a snapshot of the data Excel's 'intervals MAIN' sheet.
    Nothing existing is opened for writing, so a macro data workbook is never
    touched (its computed grades/bars stay intact).

    `data_excel` (optional) is READ (read-only) for the THICKNESS Channel/Offset
    rows and for the 'intervals MAIN' sheet copied into the output.
    Raises IntervalInputError for a bad/missing XML.
    """
    from openpyxl import Workbook, load_workbook
    from well_tools.core.xml_parser import parse_wellschematic_xml, build_pipe_summary
    from well_tools.core.intervals import build_intervals_from_xml
    from well_tools.core.thickness import parse_thickness_sections
    from well_tools.core.excel_output import _build_raw_data_sheet

    if not xml_path or not os.path.isfile(xml_path):
        raise IntervalInputError("XML file not found. Please choose a WellSchematic .xml file.")
    if not xml_path.lower().endswith(".xml"):
        raise IntervalInputError("The schematic must be a .xml file.")

    xml_data = parse_wellschematic_xml(xml_path)

    have_data = bool(data_excel and os.path.isfile(data_excel))

    # Optional THICKNESS (read-only) → Channel/Offset rows.
    thickness_sections = None
    if have_data:
        try:
            thickness_sections = parse_thickness_sections(data_excel) or None
        except ValueError:
            thickness_sections = None   # no/unreadable THICKNESS sheet — skip channels

    pipe_summary_df = build_pipe_summary(xml_data)
    interval_df = build_intervals_from_xml(xml_data, thickness_sections=thickness_sections)

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    _build_raw_data_sheet(ws, pipe_summary_df, interval_df)

    # Include a snapshot of the data Excel's 'intervals MAIN' sheet (read-only).
    copied_intervals_main = False
    if have_data:
        try:
            src = load_workbook(data_excel, data_only=True)
            match = next((s for s in src.sheetnames
                          if s.strip().lower() == INTERVALS_MAIN_SHEET.lower()), None)
            if match:
                _copy_sheet_snapshot(src[match], wb, match)
                copied_intervals_main = True
            src.close()
        except Exception:  # noqa: BLE001 — the sheet copy is best-effort
            logger.exception("Could not copy '%s' sheet", INTERVALS_MAIN_SHEET)

    wb.save(output_path)
    logger.info("Raw Data (new file): %d pipes, %d intervals, intervals_main=%s -> %s",
                len(xml_data), len(interval_df), copied_intervals_main, output_path)
    return {"output_path": output_path, "intervals_main": copied_intervals_main}
